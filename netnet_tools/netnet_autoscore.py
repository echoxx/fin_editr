#!/usr/bin/env python3
"""
Net-Net Auto-Scoring Script

Automatically scores Piotroski F-Score and C7 criteria in net-net valuation
Excel workbooks based on computed financial data.

Usage:
    python netnet_autoscore.py carmate.xlsx --price 1500
    python netnet_autoscore.py carmate.xlsx --report-only
"""

import argparse
import logging
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# Data location configuration
DATA_LOCATIONS = {
    'ncav': {
        'shares_outstanding': 20,
        'net_current_assets': 23,
        'current_ratio': 26,
        'ncav_per_share': 34,
        'ncav_qoq_change': 35,
        'ncav_yoy_change': 36,
        'debt_to_assets': 52,
        'debt_to_equity': 53,
        'liabilities_to_equity': 54,
    },
    'profitability': {
        'sales': 2,
        'gross_income': 6,
        'gross_margin': 7,
        'net_income': 10,
        'ni_margin': 11,
    },
    'ro': {
        'total_assets': 2,
        'sales': 8,
        'ebit': 9,
        'net_income': 10,
        'roa_semi': 14,
        'roa_annual': 21,
        'asset_turnover': 33,
    },
    'piotrosky': {
        'score_column': 11,  # Column K
        'notes_column': 12,  # Column L
        'total_row': 10,
    },
    'c7': {
        'score_column': 8,   # Column H
        'notes_column': 9,   # Column I
        'core_rows': (2, 10),
        'ranking_rows': (14, 22),
        'core_total_row': 11,
        'ranking_total_row': 23,
    }
}


@dataclass
class ScoreResult:
    """Result of scoring a single criterion."""
    criterion: str
    score: Optional[int]  # 0, 1, or None if can't determine
    confidence: str       # 'high', 'medium', 'low', 'manual_required'
    reasoning: str        # Explanation of how score was determined
    values_used: dict     # The actual values used in determination


class WorkbookEvaluator:
    """Evaluates cell values by tracing formulas back to raw data."""

    def __init__(self, workbook_path: str):
        self.path = workbook_path
        # Load workbook twice - once for formulas, once for raw values
        self.wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False)
        self.wb_values = openpyxl.load_workbook(workbook_path, data_only=True)

        # Find the IS and BS sheet names (case-insensitive)
        self.is_sheet = None
        self.bs_sheet = None
        for name in self.wb_formulas.sheetnames:
            lower = name.lower()
            if lower.endswith('_is'):
                self.is_sheet = name
            elif lower.endswith('_bs'):
                self.bs_sheet = name

        # Cache for evaluated values
        self._cache = {}

    def get_raw_value(self, sheet_name: str, row: int, col: int) -> Optional[float]:
        """Get raw value from IS or BS sheet, parsing string numbers."""
        ws = self.wb_formulas[sheet_name]
        val = ws.cell(row=row, column=col).value
        return self._parse_value(val)

    def _parse_value(self, val) -> Optional[float]:
        """Parse a cell value to float, handling strings and percentages."""
        if val is None or val == '' or val == 'NA' or val == '-':
            return None
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            # Remove percentage sign, commas, spaces
            cleaned = val.replace('%', '').replace(',', '').strip()
            try:
                result = float(cleaned)
                # If original had %, convert to decimal
                if '%' in val:
                    result = result / 100
                return result
            except ValueError:
                return None
        return None

    def evaluate_cell(self, sheet_name: str, row: int, col: int) -> Optional[float]:
        """Evaluate a cell, computing formula if needed."""
        cache_key = (sheet_name, row, col)
        if cache_key in self._cache:
            return self._cache[cache_key]

        ws = self.wb_formulas[sheet_name]
        cell_value = ws.cell(row=row, column=col).value

        if cell_value is None:
            return None

        # If it's a formula, evaluate it
        if isinstance(cell_value, str) and cell_value.startswith('='):
            result = self._evaluate_formula(cell_value, sheet_name)
        else:
            result = self._parse_value(cell_value)

        self._cache[cache_key] = result
        return result

    def _evaluate_formula(self, formula: str, current_sheet: str) -> Optional[float]:
        """Evaluate a formula string."""
        formula = formula[1:]  # Remove leading =

        # Handle NUMBERVALUE function
        nv_match = re.match(r'_xlfn\.NUMBERVALUE\(([^)]+)\)', formula)
        if nv_match:
            ref = nv_match.group(1)
            return self._resolve_reference(ref, current_sheet)

        # Handle simple cell references (e.g., D5)
        simple_ref = re.match(r'^([A-Z]+)(\d+)$', formula)
        if simple_ref:
            col_letter, row_num = simple_ref.groups()
            col = column_index_from_string(col_letter)
            return self.evaluate_cell(current_sheet, int(row_num), col)

        # Handle cell * constant (e.g., D14*2)
        cell_times_const = re.match(r'^([A-Z]+\d+)\*(\d+\.?\d*)$', formula)
        if cell_times_const:
            ref, const = cell_times_const.groups()
            val = self._resolve_reference(ref, current_sheet)
            if val is None:
                return None
            return val * float(const)

        # Handle cell * constant / cell (e.g., D8*4/D3)
        cell_const_div = re.match(r'^([A-Z]+\d+)\*(\d+\.?\d*)/([A-Z]+\d+)$', formula)
        if cell_const_div:
            ref1, const, ref2 = cell_const_div.groups()
            val1 = self._resolve_reference(ref1, current_sheet)
            val2 = self._resolve_reference(ref2, current_sheet)
            if val1 is None or val2 is None or val2 == 0:
                return None
            return (val1 * float(const)) / val2

        # Handle arithmetic: A-B or A/B or A*B or A+B
        arith_match = re.match(r'^([A-Z]+\d+)([-+*/])([A-Z]+\d+)$', formula)
        if arith_match:
            ref1, op, ref2 = arith_match.groups()
            val1 = self._resolve_reference(ref1, current_sheet)
            val2 = self._resolve_reference(ref2, current_sheet)
            if val1 is None or val2 is None:
                return None
            if op == '-':
                return val1 - val2
            elif op == '+':
                return val1 + val2
            elif op == '*':
                return val1 * val2
            elif op == '/':
                return val1 / val2 if val2 != 0 else None

        # Handle AVERAGE function
        avg_match = re.match(r'AVERAGE\(([^)]+)\)', formula)
        if avg_match:
            range_str = avg_match.group(1)
            return self._evaluate_average(range_str, current_sheet)

        # Couldn't evaluate
        return None

    def _resolve_reference(self, ref: str, current_sheet: str) -> Optional[float]:
        """Resolve a cell reference like 'Sheet!A1' or 'A1'."""
        if '!' in ref:
            sheet_name, cell_ref = ref.split('!')
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = current_sheet
            cell_ref = ref

        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref)
        if match:
            col_letter, row_num = match.groups()
            col = column_index_from_string(col_letter)
            return self.evaluate_cell(sheet_name, int(row_num), col)
        return None

    def _evaluate_average(self, range_str: str, current_sheet: str) -> Optional[float]:
        """Evaluate AVERAGE function."""
        # Handle range like C2:D2
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str)
        if match:
            col1, row1, col2, row2 = match.groups()
            col1_idx = column_index_from_string(col1)
            col2_idx = column_index_from_string(col2)
            row1, row2 = int(row1), int(row2)

            values = []
            for r in range(row1, row2 + 1):
                for c in range(col1_idx, col2_idx + 1):
                    val = self.evaluate_cell(current_sheet, r, c)
                    if val is not None:
                        values.append(val)

            return sum(values) / len(values) if values else None
        return None

    def get_row_series(self, sheet_name: str, row_num: int,
                       start_col: int = 4, max_cols: int = 45) -> list:
        """
        Extract values from a row, handling None/empty cells.
        Returns list of (column_index, value) tuples for non-None values.
        """
        series = []
        consecutive_empty = 0

        for col in range(start_col, start_col + max_cols):
            val = self.evaluate_cell(sheet_name, row_num, col)
            if val is not None:
                series.append((col, val))
                consecutive_empty = 0
            else:
                consecutive_empty += 1
                # Stop after 3 consecutive empty cells (after finding some data)
                if consecutive_empty >= 3 and series:
                    break

        return series

    def get_latest_n_values(self, series: list, n: int = 2) -> Optional[list]:
        """Get the N most recent values from a series."""
        if len(series) < n:
            return None
        return [v for (col, v) in series[-n:]]

    def get_latest_value(self, series: list) -> Optional[float]:
        """Get the most recent value from a series."""
        if not series:
            return None
        return series[-1][1]


class NetNetScorer:
    """Scores Piotroski F-Score and C7 criteria."""

    def __init__(self, evaluator: WorkbookEvaluator, price: Optional[float] = None,
                 burn_threshold: float = -0.10):
        self.eval = evaluator
        self.price = price
        self.burn_threshold = burn_threshold

    # =========================================================================
    # Piotroski F-Score Criteria (9 total)
    # =========================================================================

    def score_piotrosky_1_positive_earnings(self) -> ScoreResult:
        """Criterion 1: Positive net income."""
        series = self.eval.get_row_series('profitability', DATA_LOCATIONS['profitability']['net_income'], start_col=3)
        latest = self.eval.get_latest_value(series)

        if latest is None:
            return ScoreResult(
                criterion="Positive earnings",
                score=None,
                confidence="manual_required",
                reasoning="Could not find net income data",
                values_used={}
            )

        score = 1 if latest > 0 else 0
        return ScoreResult(
            criterion="Positive earnings",
            score=score,
            confidence="high",
            reasoning=f"NI = {latest:,.0f} ({'positive' if score else 'negative/zero'})",
            values_used={"net_income": latest}
        )

    def score_piotrosky_2_positive_ocf(self) -> ScoreResult:
        """Criterion 2: Positive operating cash flow."""
        # OCF typically not available in this data - flag for manual review
        return ScoreResult(
            criterion="Positive OCF",
            score=None,
            confidence="manual_required",
            reasoning="OCF data not available in standard export",
            values_used={}
        )

    def score_piotrosky_3_increasing_roa(self) -> ScoreResult:
        """Criterion 3: Increasing ROA."""
        series = self.eval.get_row_series('ro', DATA_LOCATIONS['ro']['roa_annual'])

        if len(series) < 2:
            return ScoreResult(
                criterion="Increasing ROA",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient ROA data for trend",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Increasing ROA",
                score=None,
                confidence="manual_required",
                reasoning="Could not get ROA values",
                values_used={}
            )

        increasing = values[-1] > values[-2]
        return ScoreResult(
            criterion="Increasing ROA",
            score=1 if increasing else 0,
            confidence="high",
            reasoning=f"ROA: {values[-2]:.2%} -> {values[-1]:.2%} ({'↑' if increasing else '↓'})",
            values_used={"roa_prior": values[-2], "roa_current": values[-1]}
        )

    def score_piotrosky_4_ocf_greater_ni(self) -> ScoreResult:
        """Criterion 4: OCF > Net Income (accruals quality)."""
        return ScoreResult(
            criterion="OCF > NI (accruals)",
            score=None,
            confidence="manual_required",
            reasoning="OCF data not available for accruals check",
            values_used={}
        )

    def score_piotrosky_5_decreasing_debt(self) -> ScoreResult:
        """Criterion 5: Decreasing long-term debt ratio."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['debt_to_assets'])

        if len(series) < 2:
            return ScoreResult(
                criterion="Decreasing LT debt ratio",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient debt/assets data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Decreasing LT debt ratio",
                score=None,
                confidence="manual_required",
                reasoning="Could not get debt ratio values",
                values_used={}
            )

        decreasing = values[-1] < values[-2]
        return ScoreResult(
            criterion="Decreasing LT debt ratio",
            score=1 if decreasing else 0,
            confidence="high",
            reasoning=f"D/A: {values[-2]:.2%} -> {values[-1]:.2%} ({'↓' if decreasing else '↑'})",
            values_used={"debt_ratio_prior": values[-2], "debt_ratio_current": values[-1]}
        )

    def score_piotrosky_6_increasing_cr(self) -> ScoreResult:
        """Criterion 6: Increasing current ratio."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['current_ratio'])

        if len(series) < 2:
            return ScoreResult(
                criterion="Increasing current ratio",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient current ratio data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Increasing current ratio",
                score=None,
                confidence="manual_required",
                reasoning="Could not get current ratio values",
                values_used={}
            )

        increasing = values[-1] > values[-2]
        return ScoreResult(
            criterion="Increasing current ratio",
            score=1 if increasing else 0,
            confidence="high",
            reasoning=f"CR: {values[-2]:.2f}x -> {values[-1]:.2f}x ({'↑' if increasing else '↓'})",
            values_used={"cr_prior": values[-2], "cr_current": values[-1]}
        )

    def score_piotrosky_7_no_dilution(self) -> ScoreResult:
        """Criterion 7: No share dilution (stable or decreasing shares)."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['shares_outstanding'])

        if len(series) < 2:
            return ScoreResult(
                criterion="No share dilution",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient shares data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="No share dilution",
                score=None,
                confidence="manual_required",
                reasoning="Could not get shares values",
                values_used={}
            )

        stable_or_decreasing = values[-1] <= values[-2]
        return ScoreResult(
            criterion="No share dilution",
            score=1 if stable_or_decreasing else 0,
            confidence="high",
            reasoning=f"Shares: {values[-2]:,.0f} -> {values[-1]:,.0f} ({'stable/↓' if stable_or_decreasing else '↑ dilution'})",
            values_used={"shares_prior": values[-2], "shares_current": values[-1]}
        )

    def score_piotrosky_8_increasing_gm(self) -> ScoreResult:
        """Criterion 8: Increasing gross margin."""
        series = self.eval.get_row_series('profitability', DATA_LOCATIONS['profitability']['gross_margin'], start_col=3)

        if len(series) < 2:
            return ScoreResult(
                criterion="Increasing gross margin",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient gross margin data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Increasing gross margin",
                score=None,
                confidence="manual_required",
                reasoning="Could not get gross margin values",
                values_used={}
            )

        increasing = values[-1] > values[-2]
        return ScoreResult(
            criterion="Increasing gross margin",
            score=1 if increasing else 0,
            confidence="high",
            reasoning=f"GM: {values[-2]:.1%} -> {values[-1]:.1%} ({'↑' if increasing else '↓'})",
            values_used={"gm_prior": values[-2], "gm_current": values[-1]}
        )

    def score_piotrosky_9_increasing_at(self) -> ScoreResult:
        """Criterion 9: Increasing asset turnover."""
        series = self.eval.get_row_series('ro', DATA_LOCATIONS['ro']['asset_turnover'])

        if len(series) < 2:
            return ScoreResult(
                criterion="Increasing asset turnover",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient asset turnover data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Increasing asset turnover",
                score=None,
                confidence="manual_required",
                reasoning="Could not get asset turnover values",
                values_used={}
            )

        increasing = values[-1] > values[-2]
        return ScoreResult(
            criterion="Increasing asset turnover",
            score=1 if increasing else 0,
            confidence="high",
            reasoning=f"AT: {values[-2]:.2f}x -> {values[-1]:.2f}x ({'↑' if increasing else '↓'})",
            values_used={"at_prior": values[-2], "at_current": values[-1]}
        )

    def score_all_piotrosky(self) -> list[ScoreResult]:
        """Score all Piotroski criteria.

        Note: Order matches workbook row layout (asset turnover row 8, gross margin row 9)
        """
        return [
            self.score_piotrosky_1_positive_earnings(),
            self.score_piotrosky_2_positive_ocf(),
            self.score_piotrosky_3_increasing_roa(),
            self.score_piotrosky_4_ocf_greater_ni(),
            self.score_piotrosky_5_decreasing_debt(),
            self.score_piotrosky_6_increasing_cr(),
            self.score_piotrosky_7_no_dilution(),
            self.score_piotrosky_9_increasing_at(),  # Row 8 in workbook
            self.score_piotrosky_8_increasing_gm(),  # Row 9 in workbook
        ]

    # =========================================================================
    # C7 Core Criteria (9 total)
    # =========================================================================

    def score_c7_core_1_not_chinese(self) -> ScoreResult:
        """Core 1: Not majority Chinese owned."""
        # Check Overview sheet for country if it exists
        return ScoreResult(
            criterion="Not Majority Chinese",
            score=None,
            confidence="manual_required",
            reasoning="Check company domicile manually",
            values_used={}
        )

    def score_c7_core_2_low_price_ncav(self) -> ScoreResult:
        """Core 2: Low Price-to-NCAV (< 0.67)."""
        if self.price is None:
            return ScoreResult(
                criterion="Low P/NCAV",
                score=None,
                confidence="manual_required",
                reasoning="Price not provided",
                values_used={}
            )

        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['ncav_per_share'])
        ncav = self.eval.get_latest_value(series)

        if ncav is None or ncav <= 0:
            return ScoreResult(
                criterion="Low P/NCAV",
                score=None,
                confidence="manual_required",
                reasoning=f"NCAV not available or non-positive: {ncav}",
                values_used={}
            )

        p_ncav = self.price / ncav
        score = 1 if p_ncav < 0.67 else 0
        return ScoreResult(
            criterion="Low P/NCAV",
            score=score,
            confidence="high",
            reasoning=f"P/NCAV = {p_ncav:.2f} (threshold: 0.67)",
            values_used={"price": self.price, "ncav": ncav, "p_ncav": p_ncav}
        )

    def score_c7_core_3_low_debt_equity(self) -> ScoreResult:
        """Core 3: Low Debt-to-Equity (< 0.5)."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['debt_to_equity'])
        de = self.eval.get_latest_value(series)

        if de is None:
            return ScoreResult(
                criterion="Low D/E",
                score=None,
                confidence="manual_required",
                reasoning="D/E data not available",
                values_used={}
            )

        score = 1 if de < 0.5 else 0
        return ScoreResult(
            criterion="Low D/E",
            score=score,
            confidence="high",
            reasoning=f"D/E = {de:.2f} (threshold: 0.5)",
            values_used={"debt_to_equity": de}
        )

    def score_c7_core_4_adequate_earnings(self) -> ScoreResult:
        """Core 4: Adequate past earnings (any period with NI margin > 5%)."""
        series = self.eval.get_row_series('profitability', DATA_LOCATIONS['profitability']['ni_margin'], start_col=3)

        if not series:
            return ScoreResult(
                criterion="Adequate past earnings",
                score=None,
                confidence="manual_required",
                reasoning="NI margin data not available",
                values_used={}
            )

        # Check if any period has > 5% margin
        has_adequate = any(v > 0.05 for (_, v) in series)
        max_margin = max(v for (_, v) in series)

        return ScoreResult(
            criterion="Adequate past earnings",
            score=1 if has_adequate else 0,
            confidence="high",
            reasoning=f"Max NI margin = {max_margin:.1%} (threshold: 5%)",
            values_used={"max_ni_margin": max_margin, "periods_checked": len(series)}
        )

    def score_c7_core_5_past_price_above_ncav(self) -> ScoreResult:
        """Core 5: Past price above NCAV."""
        return ScoreResult(
            criterion="Past price above NCAV",
            score=None,
            confidence="manual_required",
            reasoning="Requires historical price data",
            values_used={}
        )

    def score_c7_core_6_existing_operations(self) -> ScoreResult:
        """Core 6: Existing operations (not shell company)."""
        return ScoreResult(
            criterion="Existing operations",
            score=None,
            confidence="manual_required",
            reasoning="Qualitative assessment required",
            values_used={}
        )

    def score_c7_core_7_not_selling_shares(self) -> ScoreResult:
        """Core 7: Not selling shares (stable or decreasing)."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['shares_outstanding'])

        if len(series) < 2:
            return ScoreResult(
                criterion="Not selling shares",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient shares data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Not selling shares",
                score=None,
                confidence="manual_required",
                reasoning="Could not get shares values",
                values_used={}
            )

        not_selling = values[-1] <= values[-2]
        return ScoreResult(
            criterion="Not selling shares",
            score=1 if not_selling else 0,
            confidence="high",
            reasoning=f"Shares: {values[-2]:,.0f} -> {values[-1]:,.0f}",
            values_used={"shares_prior": values[-2], "shares_current": values[-1]}
        )

    def score_c7_core_8_small_market_cap(self) -> ScoreResult:
        """Core 8: Market cap < $50mm."""
        if self.price is None:
            return ScoreResult(
                criterion="Market cap < $50mm",
                score=None,
                confidence="manual_required",
                reasoning="Price not provided",
                values_used={}
            )

        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['shares_outstanding'])
        shares = self.eval.get_latest_value(series)

        if shares is None:
            return ScoreResult(
                criterion="Market cap < $50mm",
                score=None,
                confidence="manual_required",
                reasoning="Shares data not available",
                values_used={}
            )

        # Note: price and shares may be in different units (shares in millions, price in yen, etc.)
        # Market cap calculation depends on the currency and units
        market_cap = self.price * shares

        # For Japanese stocks, this will be in yen. 50M USD ≈ 7.5B yen
        # Flag for manual review since currency varies
        return ScoreResult(
            criterion="Market cap < $50mm",
            score=None,
            confidence="manual_required",
            reasoning=f"Market cap = {market_cap:,.0f} (currency/unit check needed)",
            values_used={"price": self.price, "shares": shares, "market_cap": market_cap}
        )

    def score_c7_core_9_low_burn_rate(self) -> ScoreResult:
        """Core 9: Low NCAV burn rate (YoY decline < threshold)."""
        # Try to get pre-calculated YoY change
        yoy_series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['ncav_yoy_change'])

        if yoy_series:
            latest_yoy = self.eval.get_latest_value(yoy_series)
            if latest_yoy is not None:
                score = 1 if latest_yoy > self.burn_threshold else 0
                return ScoreResult(
                    criterion="Low NCAV burn rate",
                    score=score,
                    confidence="high",
                    reasoning=f"NCAV YoY: {latest_yoy:.1%} (threshold: {self.burn_threshold:.0%})",
                    values_used={"ncav_yoy_change": latest_yoy}
                )

        # Fallback: calculate from NCAV series (need 3 periods for YoY in semi-annual)
        ncav_series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['ncav_per_share'])

        if len(ncav_series) < 3:
            return ScoreResult(
                criterion="Low NCAV burn rate",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient NCAV history for YoY calculation",
                values_used={}
            )

        values = [v for (_, v) in ncav_series[-3:]]
        if values[-3] == 0:
            return ScoreResult(
                criterion="Low NCAV burn rate",
                score=None,
                confidence="manual_required",
                reasoning="Cannot calculate YoY (zero denominator)",
                values_used={}
            )

        yoy_change = (values[-1] - values[-3]) / abs(values[-3])
        score = 1 if yoy_change > self.burn_threshold else 0

        return ScoreResult(
            criterion="Low NCAV burn rate",
            score=score,
            confidence="medium",
            reasoning=f"Calculated NCAV YoY: {yoy_change:.1%}",
            values_used={"ncav_yoy_calculated": yoy_change}
        )

    def score_all_c7_core(self) -> list[ScoreResult]:
        """Score all C7 core criteria."""
        return [
            self.score_c7_core_1_not_chinese(),
            self.score_c7_core_2_low_price_ncav(),
            self.score_c7_core_3_low_debt_equity(),
            self.score_c7_core_4_adequate_earnings(),
            self.score_c7_core_5_past_price_above_ncav(),
            self.score_c7_core_6_existing_operations(),
            self.score_c7_core_7_not_selling_shares(),
            self.score_c7_core_8_small_market_cap(),
            self.score_c7_core_9_low_burn_rate(),
        ]

    # =========================================================================
    # C7 Ranking Criteria (9 total)
    # =========================================================================

    def score_c7_rank_1_high_current_ratio(self) -> ScoreResult:
        """Ranking 1: Current ratio > 1.5x."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['current_ratio'])
        cr = self.eval.get_latest_value(series)

        if cr is None:
            return ScoreResult(
                criterion="CR > 1.5x",
                score=None,
                confidence="manual_required",
                reasoning="Current ratio not available",
                values_used={}
            )

        score = 1 if cr > 1.5 else 0
        return ScoreResult(
            criterion="CR > 1.5x",
            score=score,
            confidence="high",
            reasoning=f"CR = {cr:.2f}x (threshold: 1.5x)",
            values_used={"current_ratio": cr}
        )

    def score_c7_rank_2_not_financial(self) -> ScoreResult:
        """Ranking 2: Not financial/real estate/fund."""
        return ScoreResult(
            criterion="Not financial/RE/fund",
            score=None,
            confidence="manual_required",
            reasoning="Check industry classification manually",
            values_used={}
        )

    def score_c7_rank_3_buying_back(self) -> ScoreResult:
        """Ranking 3: Company is buying back stock (shares decreasing)."""
        series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['shares_outstanding'])

        if len(series) < 2:
            return ScoreResult(
                criterion="Buying back stock",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient shares data",
                values_used={}
            )

        values = self.eval.get_latest_n_values(series, 2)
        if values is None:
            return ScoreResult(
                criterion="Buying back stock",
                score=None,
                confidence="manual_required",
                reasoning="Could not get shares values",
                values_used={}
            )

        buying_back = values[-1] < values[-2]
        return ScoreResult(
            criterion="Buying back stock",
            score=1 if buying_back else 0,
            confidence="high",
            reasoning=f"Shares: {values[-2]:,.0f} -> {values[-1]:,.0f} ({'↓ buyback' if buying_back else 'no buyback'})",
            values_used={"shares_prior": values[-2], "shares_current": values[-1]}
        )

    def score_c7_rank_4_low_price_net_cash(self) -> ScoreResult:
        """Ranking 4: Low price-to-net cash."""
        return ScoreResult(
            criterion="Low P/Net Cash",
            score=None,
            confidence="manual_required",
            reasoning="Net cash calculation requires manual review",
            values_used={}
        )

    def score_c7_rank_5_insider_ownership(self) -> ScoreResult:
        """Ranking 5: Insider ownership > 10%."""
        return ScoreResult(
            criterion="Insider ownership > 10%",
            score=None,
            confidence="manual_required",
            reasoning="Check ownership sheet manually",
            values_used={}
        )

    def score_c7_rank_6_insider_buys(self) -> ScoreResult:
        """Ranking 6: Insider buys > sells."""
        return ScoreResult(
            criterion="Insider buys > sells",
            score=None,
            confidence="manual_required",
            reasoning="Requires insider transaction data",
            values_used={}
        )

    def score_c7_rank_7_positive_burn(self) -> ScoreResult:
        """Ranking 7: Positive burn rate (NCAV growing)."""
        # Try pre-calculated YoY
        yoy_series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['ncav_yoy_change'])

        if yoy_series:
            latest_yoy = self.eval.get_latest_value(yoy_series)
            if latest_yoy is not None:
                score = 1 if latest_yoy > 0 else 0
                return ScoreResult(
                    criterion="Positive burn rate",
                    score=score,
                    confidence="high",
                    reasoning=f"NCAV YoY: {latest_yoy:.1%} ({'growing' if score else 'declining'})",
                    values_used={"ncav_yoy_change": latest_yoy}
                )

        # Fallback calculation
        ncav_series = self.eval.get_row_series('ncav', DATA_LOCATIONS['ncav']['ncav_per_share'])

        if len(ncav_series) < 3:
            return ScoreResult(
                criterion="Positive burn rate",
                score=None,
                confidence="manual_required",
                reasoning="Insufficient NCAV history",
                values_used={}
            )

        values = [v for (_, v) in ncav_series[-3:]]
        if values[-3] == 0:
            return ScoreResult(
                criterion="Positive burn rate",
                score=None,
                confidence="manual_required",
                reasoning="Cannot calculate (zero denominator)",
                values_used={}
            )

        yoy_change = (values[-1] - values[-3]) / abs(values[-3])
        score = 1 if yoy_change > 0 else 0

        return ScoreResult(
            criterion="Positive burn rate",
            score=score,
            confidence="medium",
            reasoning=f"NCAV YoY: {yoy_change:.1%}",
            values_used={"ncav_yoy_calculated": yoy_change}
        )

    def score_c7_rank_8_reasonable_pay(self) -> ScoreResult:
        """Ranking 8: Reasonable insider pay."""
        return ScoreResult(
            criterion="Reasonable insider pay",
            score=None,
            confidence="manual_required",
            reasoning="Qualitative assessment required",
            values_used={}
        )

    def score_c7_rank_9_dividend_yield(self) -> ScoreResult:
        """Ranking 9: Dividend yield."""
        return ScoreResult(
            criterion="Dividend yield",
            score=None,
            confidence="manual_required",
            reasoning="Requires dividend and price data",
            values_used={}
        )

    def score_all_c7_ranking(self) -> list[ScoreResult]:
        """Score all C7 ranking criteria."""
        return [
            self.score_c7_rank_1_high_current_ratio(),
            self.score_c7_rank_2_not_financial(),
            self.score_c7_rank_3_buying_back(),
            self.score_c7_rank_4_low_price_net_cash(),
            self.score_c7_rank_5_insider_ownership(),
            self.score_c7_rank_6_insider_buys(),
            self.score_c7_rank_7_positive_burn(),
            self.score_c7_rank_8_reasonable_pay(),
            self.score_c7_rank_9_dividend_yield(),
        ]


def print_report(piotrosky_results: list[ScoreResult],
                 c7_core_results: list[ScoreResult],
                 c7_rank_results: list[ScoreResult]):
    """Print a formatted scoring report."""

    def print_section(title: str, results: list[ScoreResult]):
        print(f"\n{'=' * 60}")
        print(f"{title}")
        print('=' * 60)

        auto_count = 0
        manual_count = 0
        total_score = 0

        for i, r in enumerate(results, 1):
            if r.score is not None:
                auto_count += 1
                total_score += r.score
                score_str = str(r.score)
                status = "AUTO"
            else:
                manual_count += 1
                score_str = "-"
                status = "MANUAL"

            print(f"{i}. {r.criterion:30s} {score_str:>3s}  ({status}) {r.reasoning}")

        print('-' * 60)
        print(f"Auto-scored: {auto_count}/{len(results)} | Manual required: {manual_count}/{len(results)} | Score: {total_score}/{auto_count if auto_count else '?'}")

    print_section("PIOTROSKI F-SCORE", piotrosky_results)
    print_section("C7 CORE CRITERIA", c7_core_results)
    print_section("C7 RANKING CRITERIA", c7_rank_results)


def write_scores_to_workbook(wb: openpyxl.Workbook,
                             piotrosky_results: list[ScoreResult],
                             c7_core_results: list[ScoreResult],
                             c7_rank_results: list[ScoreResult]):
    """Write scores and notes to the workbook."""

    # Write Piotroski scores
    ws = wb['piotrosky']
    score_col = DATA_LOCATIONS['piotrosky']['score_column']
    notes_col = DATA_LOCATIONS['piotrosky']['notes_column']

    for i, r in enumerate(piotrosky_results):
        row = i + 1  # Rows 1-9
        if r.score is not None:
            ws.cell(row=row, column=score_col).value = r.score
            ws.cell(row=row, column=notes_col).value = f"AUTO: {r.reasoning}"
        else:
            ws.cell(row=row, column=notes_col).value = f"MANUAL: {r.reasoning}"

    # Write C7 Core scores
    ws = wb['C7']
    score_col = DATA_LOCATIONS['c7']['score_column']
    notes_col = DATA_LOCATIONS['c7']['notes_column']

    for i, r in enumerate(c7_core_results):
        row = i + 2  # Rows 2-10
        if r.score is not None:
            ws.cell(row=row, column=score_col).value = r.score
            ws.cell(row=row, column=notes_col).value = f"AUTO: {r.reasoning}"
        else:
            ws.cell(row=row, column=notes_col).value = f"MANUAL: {r.reasoning}"

    # Write C7 Ranking scores
    for i, r in enumerate(c7_rank_results):
        row = i + 14  # Rows 14-22
        if r.score is not None:
            ws.cell(row=row, column=score_col).value = r.score
            ws.cell(row=row, column=notes_col).value = f"AUTO: {r.reasoning}"
        else:
            ws.cell(row=row, column=notes_col).value = f"MANUAL: {r.reasoning}"


def create_backup(filepath: str) -> str:
    """Create a timestamped backup of the workbook."""
    path = Path(filepath)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"{path.stem}_backup_{timestamp}{path.suffix}"
    backup_path = path.parent / backup_name
    shutil.copy2(filepath, backup_path)
    return str(backup_path)


def main():
    parser = argparse.ArgumentParser(
        description='Auto-score net-net Piotroski and C7 criteria'
    )
    parser.add_argument('workbook', help='Path to the Excel workbook')
    parser.add_argument('--price', type=float, help='Current share price for P/NCAV calculations')
    parser.add_argument('--output', help='Output path (default: modify in place)')
    parser.add_argument('--report-only', action='store_true',
                        help='Print report without modifying workbook')
    parser.add_argument('--burn-threshold', type=float, default=-0.10,
                        help='NCAV burn rate threshold (default: -0.10)')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Show detailed debug output')

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Load and evaluate workbook
    logger.info(f"Loading workbook: {args.workbook}")
    evaluator = WorkbookEvaluator(args.workbook)

    if evaluator.is_sheet:
        logger.info(f"Found IS sheet: {evaluator.is_sheet}")
    if evaluator.bs_sheet:
        logger.info(f"Found BS sheet: {evaluator.bs_sheet}")

    # Score all criteria
    scorer = NetNetScorer(evaluator, price=args.price, burn_threshold=args.burn_threshold)

    logger.info("Scoring Piotroski criteria...")
    piotrosky_results = scorer.score_all_piotrosky()

    logger.info("Scoring C7 Core criteria...")
    c7_core_results = scorer.score_all_c7_core()

    logger.info("Scoring C7 Ranking criteria...")
    c7_rank_results = scorer.score_all_c7_ranking()

    # Print report
    print_report(piotrosky_results, c7_core_results, c7_rank_results)

    # Write to workbook if not report-only
    if not args.report_only:
        # Create backup
        backup_path = create_backup(args.workbook)
        logger.info(f"Backup created: {backup_path}")

        # Write scores
        write_scores_to_workbook(
            evaluator.wb_formulas,
            piotrosky_results,
            c7_core_results,
            c7_rank_results
        )

        # Save
        output_path = args.output or args.workbook
        evaluator.wb_formulas.save(output_path)
        logger.info(f"Scores written to: {output_path}")
    else:
        logger.info("Report-only mode - workbook not modified")


if __name__ == '__main__':
    main()
