#!/usr/bin/env python3
"""
netnet_validator.py - Validates and maps structure of net-net analysis workbooks

This tool:
1. Maps the structure of an existing workbook to a JSON file
2. Validates new Investing.com exports against an existing structure
3. Generates diff reports showing what would change if new data were applied

Usage:
    # Map structure from an existing workbook
    python netnet_validator.py map --workbook almedio.xlsx --output almedio_structure.json

    # Validate new data against existing structure
    python netnet_validator.py validate --structure almedio_structure.json \
        --is-file "Almedio Inc - Income Statement.xlsx" \
        --bs-file "Almedio Inc - Balance Sheet.xlsx"

    # Generate diff report
    python netnet_validator.py diff --workbook almedio.xlsx \
        --is-file "Almedio Inc - Income Statement.xlsx" \
        --bs-file "Almedio Inc - Balance Sheet.xlsx"
"""

import argparse
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class WorkbookStructure:
    """Represents the structure of a net-net analysis workbook."""

    def __init__(self):
        self.company_name: str = ""
        self.is_sheet_name: str = ""
        self.bs_sheet_name: str = ""
        self.is_row_labels: dict[int, str] = {}  # row_number -> label
        self.bs_row_labels: dict[int, str] = {}
        self.is_column_dates: dict[int, str] = {}  # col_number -> date_string
        self.bs_column_dates: dict[int, str] = {}
        self.is_period_codes: dict[int, str] = {}  # col_number -> period_code
        self.bs_period_codes: dict[int, str] = {}
        self.extraction_date: str = ""
        self.source_file: str = ""

    def to_dict(self) -> dict:
        """Convert structure to dictionary for JSON serialization."""
        return {
            "company_name": self.company_name,
            "is_sheet_name": self.is_sheet_name,
            "bs_sheet_name": self.bs_sheet_name,
            "income_statement": {
                "row_labels": {str(k): v for k, v in self.is_row_labels.items()},
                "column_dates": {str(k): v for k, v in self.is_column_dates.items()},
                "period_codes": {str(k): v for k, v in self.is_period_codes.items()}
            },
            "balance_sheet": {
                "row_labels": {str(k): v for k, v in self.bs_row_labels.items()},
                "column_dates": {str(k): v for k, v in self.bs_column_dates.items()},
                "period_codes": {str(k): v for k, v in self.bs_period_codes.items()}
            },
            "metadata": {
                "extraction_date": self.extraction_date,
                "source_file": self.source_file
            }
        }

    @classmethod
    def from_dict(cls, data: dict) -> "WorkbookStructure":
        """Create structure from dictionary (loaded from JSON)."""
        structure = cls()
        structure.company_name = data.get("company_name", "")
        structure.is_sheet_name = data.get("is_sheet_name", "")
        structure.bs_sheet_name = data.get("bs_sheet_name", "")

        is_data = data.get("income_statement", {})
        structure.is_row_labels = {int(k): v for k, v in is_data.get("row_labels", {}).items()}
        structure.is_column_dates = {int(k): v for k, v in is_data.get("column_dates", {}).items()}
        structure.is_period_codes = {int(k): v for k, v in is_data.get("period_codes", {}).items()}

        bs_data = data.get("balance_sheet", {})
        structure.bs_row_labels = {int(k): v for k, v in bs_data.get("row_labels", {}).items()}
        structure.bs_column_dates = {int(k): v for k, v in bs_data.get("column_dates", {}).items()}
        structure.bs_period_codes = {int(k): v for k, v in bs_data.get("period_codes", {}).items()}

        metadata = data.get("metadata", {})
        structure.extraction_date = metadata.get("extraction_date", "")
        structure.source_file = metadata.get("source_file", "")

        return structure


def find_sheets(workbook: openpyxl.Workbook) -> tuple[str | None, str | None]:
    """Find the Income Statement and Balance Sheet sheets by naming convention."""
    is_sheet = None
    bs_sheet = None

    for name in workbook.sheetnames:
        if name.endswith('_IS'):
            is_sheet = name
        elif name.endswith('_bs'):
            bs_sheet = name

    return is_sheet, bs_sheet


def extract_row_labels(worksheet, start_row: int = 12, max_row: int = 60, label_col: int = 3) -> dict[int, str]:
    """Extract row labels from a worksheet."""
    labels = {}
    for row in range(start_row, max_row + 1):
        value = worksheet.cell(row=row, column=label_col).value
        if value and isinstance(value, str) and value.strip():
            labels[row] = value.strip()
    return labels


def extract_column_dates(worksheet, date_row: int = 10, start_col: int = 4, max_col: int = 50) -> dict[int, str]:
    """Extract column dates from a worksheet."""
    dates = {}
    for col in range(start_col, max_col + 1):
        value = worksheet.cell(row=date_row, column=col).value
        if value:
            # Handle datetime objects or strings
            if hasattr(value, 'strftime'):
                dates[col] = value.strftime('%Y-%m-%d')
            elif isinstance(value, str) and value.strip():
                dates[col] = value.strip()
    return dates


def extract_period_codes(worksheet, code_row: int = 8, start_col: int = 4, max_col: int = 50) -> dict[int, str]:
    """Extract period codes from a worksheet."""
    codes = {}
    for col in range(start_col, max_col + 1):
        value = worksheet.cell(row=code_row, column=col).value
        if value and isinstance(value, str) and value.strip():
            codes[col] = value.strip()
    return codes


def map_workbook_structure(workbook_path: str) -> WorkbookStructure:
    """Map the complete structure of a workbook to a WorkbookStructure object."""
    logger.info(f"Mapping structure from: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    structure = WorkbookStructure()
    structure.source_file = str(Path(workbook_path).name)
    structure.extraction_date = datetime.now().isoformat()

    # Find IS and BS sheets
    is_sheet, bs_sheet = find_sheets(wb)

    if not is_sheet:
        raise ValueError("Could not find Income Statement sheet (ending in '_IS')")
    if not bs_sheet:
        raise ValueError("Could not find Balance Sheet sheet (ending in '_bs')")

    structure.is_sheet_name = is_sheet
    structure.bs_sheet_name = bs_sheet

    # Extract company name from IS sheet
    is_ws = wb[is_sheet]
    structure.company_name = is_ws['C2'].value or ""

    # Extract Income Statement structure
    structure.is_row_labels = extract_row_labels(is_ws)
    structure.is_column_dates = extract_column_dates(is_ws)
    structure.is_period_codes = extract_period_codes(is_ws)

    # Extract Balance Sheet structure
    bs_ws = wb[bs_sheet]
    structure.bs_row_labels = extract_row_labels(bs_ws)
    structure.bs_column_dates = extract_column_dates(bs_ws)
    structure.bs_period_codes = extract_period_codes(bs_ws)

    logger.info(f"Mapped {len(structure.is_row_labels)} IS rows, {len(structure.bs_row_labels)} BS rows")
    logger.info(f"Found {len(structure.is_column_dates)} date columns")

    wb.close()
    return structure


def map_export_structure(export_path: str, sheet_type: str) -> tuple[dict[int, str], dict[int, str], dict[int, str]]:
    """Map the structure of an Investing.com export file.

    Returns:
        Tuple of (row_labels, column_dates, period_codes)
    """
    logger.info(f"Mapping export structure from: {export_path}")

    wb = openpyxl.load_workbook(export_path, data_only=False)
    ws = wb.active

    row_labels = extract_row_labels(ws)
    column_dates = extract_column_dates(ws)
    period_codes = extract_period_codes(ws)

    logger.info(f"Export has {len(row_labels)} rows, {len(column_dates)} date columns")

    wb.close()
    return row_labels, column_dates, period_codes


class ValidationResult:
    """Holds results from validating new data against existing structure."""

    def __init__(self, sheet_type: str):
        self.sheet_type = sheet_type
        self.is_valid = True
        self.row_mismatches: list[dict] = []  # {row, expected, actual}
        self.missing_rows: list[dict] = []  # {row, expected_label}
        self.new_rows: list[dict] = []  # {row, label}
        self.new_periods: list[dict] = []  # {col, date, period_code}
        self.removed_periods: list[dict] = []  # {col, date, period_code}
        self.warnings: list[str] = []
        self.errors: list[str] = []

    def add_error(self, msg: str):
        self.errors.append(msg)
        self.is_valid = False

    def to_dict(self) -> dict:
        return {
            "sheet_type": self.sheet_type,
            "is_valid": self.is_valid,
            "row_mismatches": self.row_mismatches,
            "missing_rows": self.missing_rows,
            "new_rows": self.new_rows,
            "new_periods": self.new_periods,
            "removed_periods": self.removed_periods,
            "warnings": self.warnings,
            "errors": self.errors
        }


def validate_export_against_structure(
    structure: WorkbookStructure,
    export_path: str,
    sheet_type: str  # 'is' or 'bs'
) -> ValidationResult:
    """Validate an Investing.com export against expected structure."""

    result = ValidationResult(sheet_type)

    if sheet_type == 'is':
        expected_labels = structure.is_row_labels
        expected_dates = structure.is_column_dates
        expected_codes = structure.is_period_codes
    else:
        expected_labels = structure.bs_row_labels
        expected_dates = structure.bs_column_dates
        expected_codes = structure.bs_period_codes

    # Map export structure
    export_labels, export_dates, export_codes = map_export_structure(export_path, sheet_type)

    # Check row label alignment
    for row, expected_label in expected_labels.items():
        if row not in export_labels:
            result.missing_rows.append({
                "row": row,
                "expected_label": expected_label
            })
            result.add_error(f"Row {row} missing in export (expected: '{expected_label}')")
        elif export_labels[row] != expected_label:
            result.row_mismatches.append({
                "row": row,
                "expected": expected_label,
                "actual": export_labels[row]
            })
            result.add_error(f"Row {row} mismatch: expected '{expected_label}', got '{export_labels[row]}'")

    # Check for new rows in export
    for row, label in export_labels.items():
        if row not in expected_labels:
            result.new_rows.append({
                "row": row,
                "label": label
            })
            result.warnings.append(f"New row {row} in export: '{label}'")

    # Check for new periods
    export_date_set = set(export_dates.values())
    expected_date_set = set(expected_dates.values())

    new_dates = export_date_set - expected_date_set
    removed_dates = expected_date_set - export_date_set

    for col, date in export_dates.items():
        if date in new_dates:
            period_code = export_codes.get(col, "")
            result.new_periods.append({
                "col": col,
                "date": date,
                "period_code": period_code
            })

    for col, date in expected_dates.items():
        if date in removed_dates:
            period_code = expected_codes.get(col, "")
            result.removed_periods.append({
                "col": col,
                "date": date,
                "period_code": period_code
            })

    if result.new_periods:
        result.warnings.append(f"Found {len(result.new_periods)} new period(s) in export")

    if result.removed_periods:
        result.warnings.append(f"Export missing {len(result.removed_periods)} period(s) from existing workbook")

    return result


class DiffReport:
    """Holds a diff report comparing existing workbook to new data."""

    def __init__(self):
        self.is_changes: list[dict] = []  # {row, col, old_value, new_value, label}
        self.bs_changes: list[dict] = []
        self.new_is_periods: list[dict] = []
        self.new_bs_periods: list[dict] = []
        self.summary: dict = {}

    def to_dict(self) -> dict:
        return {
            "income_statement_changes": self.is_changes,
            "balance_sheet_changes": self.bs_changes,
            "new_is_periods": self.new_is_periods,
            "new_bs_periods": self.new_bs_periods,
            "summary": self.summary
        }


def generate_diff_report(
    workbook_path: str,
    is_export_path: str | None = None,
    bs_export_path: str | None = None
) -> DiffReport:
    """Generate a diff report showing what would change if new data were applied."""

    report = DiffReport()

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    is_sheet, bs_sheet = find_sheets(wb)

    is_change_count = 0
    bs_change_count = 0

    if is_export_path and is_sheet:
        export_wb = openpyxl.load_workbook(is_export_path, data_only=False)
        export_ws = export_wb.active
        existing_ws = wb[is_sheet]

        # Get existing dates to find matching columns
        existing_dates = extract_column_dates(existing_ws)
        export_dates = extract_column_dates(export_ws)

        # Map export columns to existing columns by date
        date_to_existing_col = {date: col for col, date in existing_dates.items()}
        date_to_export_col = {date: col for col, date in export_dates.items()}

        # Find new periods
        for date, export_col in date_to_export_col.items():
            if date not in date_to_existing_col:
                period_code = export_ws.cell(row=8, column=export_col).value
                report.new_is_periods.append({
                    "date": date,
                    "period_code": period_code,
                    "export_col": export_col
                })

        # Compare values for matching dates
        row_labels = extract_row_labels(existing_ws)
        for date in date_to_existing_col:
            if date in date_to_export_col:
                existing_col = date_to_existing_col[date]
                export_col = date_to_export_col[date]

                for row, label in row_labels.items():
                    existing_val = existing_ws.cell(row=row, column=existing_col).value
                    new_val = export_ws.cell(row=row, column=export_col).value

                    # Normalize for comparison
                    existing_str = str(existing_val).strip() if existing_val else ""
                    new_str = str(new_val).strip() if new_val else ""

                    if existing_str != new_str:
                        report.is_changes.append({
                            "row": row,
                            "col": existing_col,
                            "col_letter": get_column_letter(existing_col),
                            "date": date,
                            "label": label,
                            "old_value": existing_val,
                            "new_value": new_val
                        })
                        is_change_count += 1

        export_wb.close()

    if bs_export_path and bs_sheet:
        export_wb = openpyxl.load_workbook(bs_export_path, data_only=False)
        export_ws = export_wb.active
        existing_ws = wb[bs_sheet]

        # Get existing dates to find matching columns
        existing_dates = extract_column_dates(existing_ws)
        export_dates = extract_column_dates(export_ws)

        # Map export columns to existing columns by date
        date_to_existing_col = {date: col for col, date in existing_dates.items()}
        date_to_export_col = {date: col for col, date in export_dates.items()}

        # Find new periods
        for date, export_col in date_to_export_col.items():
            if date not in date_to_existing_col:
                period_code = export_ws.cell(row=8, column=export_col).value
                report.new_bs_periods.append({
                    "date": date,
                    "period_code": period_code,
                    "export_col": export_col
                })

        # Compare values for matching dates
        row_labels = extract_row_labels(existing_ws)
        for date in date_to_existing_col:
            if date in date_to_export_col:
                existing_col = date_to_existing_col[date]
                export_col = date_to_export_col[date]

                for row, label in row_labels.items():
                    existing_val = existing_ws.cell(row=row, column=existing_col).value
                    new_val = export_ws.cell(row=row, column=export_col).value

                    # Normalize for comparison
                    existing_str = str(existing_val).strip() if existing_val else ""
                    new_str = str(new_val).strip() if new_val else ""

                    if existing_str != new_str:
                        report.bs_changes.append({
                            "row": row,
                            "col": existing_col,
                            "col_letter": get_column_letter(existing_col),
                            "date": date,
                            "label": label,
                            "old_value": existing_val,
                            "new_value": new_val
                        })
                        bs_change_count += 1

        export_wb.close()

    wb.close()

    report.summary = {
        "income_statement_changes": is_change_count,
        "balance_sheet_changes": bs_change_count,
        "new_is_periods": len(report.new_is_periods),
        "new_bs_periods": len(report.new_bs_periods)
    }

    return report


def cmd_map(args):
    """Handle the 'map' subcommand."""
    structure = map_workbook_structure(args.workbook)

    output_path = args.output
    if not output_path:
        # Default to company name based file
        base_name = Path(args.workbook).stem
        output_path = f"{base_name}_structure.json"

    with open(output_path, 'w') as f:
        json.dump(structure.to_dict(), f, indent=2)

    logger.info(f"Structure saved to: {output_path}")

    # Print summary
    print(f"\nWorkbook Structure Summary")
    print(f"=" * 50)
    print(f"Company: {structure.company_name}")
    print(f"IS Sheet: {structure.is_sheet_name}")
    print(f"BS Sheet: {structure.bs_sheet_name}")
    print(f"\nIncome Statement:")
    print(f"  Row labels: {len(structure.is_row_labels)}")
    print(f"  Date columns: {len(structure.is_column_dates)}")
    print(f"\nBalance Sheet:")
    print(f"  Row labels: {len(structure.bs_row_labels)}")
    print(f"  Date columns: {len(structure.bs_column_dates)}")


def cmd_validate(args):
    """Handle the 'validate' subcommand."""
    # Load structure
    with open(args.structure, 'r') as f:
        data = json.load(f)
    structure = WorkbookStructure.from_dict(data)

    results = []

    if args.is_file:
        is_result = validate_export_against_structure(structure, args.is_file, 'is')
        results.append(is_result)

    if args.bs_file:
        bs_result = validate_export_against_structure(structure, args.bs_file, 'bs')
        results.append(bs_result)

    # Print results
    all_valid = True
    for result in results:
        sheet_name = "Income Statement" if result.sheet_type == 'is' else "Balance Sheet"
        print(f"\n{sheet_name} Validation")
        print("=" * 50)

        if result.is_valid:
            print("Status: VALID")
        else:
            print("Status: INVALID")
            all_valid = False

        if result.errors:
            print("\nErrors:")
            for error in result.errors:
                print(f"  - {error}")

        if result.warnings:
            print("\nWarnings:")
            for warning in result.warnings:
                print(f"  - {warning}")

        if result.new_periods:
            print("\nNew periods available:")
            for period in result.new_periods:
                print(f"  - {period['date']} ({period['period_code']})")

    if args.output:
        output_data = {
            "validation_date": datetime.now().isoformat(),
            "results": [r.to_dict() for r in results]
        }
        with open(args.output, 'w') as f:
            json.dump(output_data, f, indent=2)
        logger.info(f"Validation results saved to: {args.output}")

    return 0 if all_valid else 1


def cmd_diff(args):
    """Handle the 'diff' subcommand."""
    report = generate_diff_report(
        args.workbook,
        args.is_file,
        args.bs_file
    )

    # Print summary
    print(f"\nDiff Report Summary")
    print("=" * 50)
    print(f"Income Statement changes: {report.summary['income_statement_changes']}")
    print(f"Balance Sheet changes: {report.summary['balance_sheet_changes']}")
    print(f"New IS periods: {report.summary['new_is_periods']}")
    print(f"New BS periods: {report.summary['new_bs_periods']}")

    if report.new_is_periods:
        print("\nNew Income Statement periods:")
        for period in report.new_is_periods:
            print(f"  - {period['date']} ({period['period_code']})")

    if report.new_bs_periods:
        print("\nNew Balance Sheet periods:")
        for period in report.new_bs_periods:
            print(f"  - {period['date']} ({period['period_code']})")

    if args.verbose:
        if report.is_changes:
            print(f"\nIncome Statement value changes ({len(report.is_changes)}):")
            for change in report.is_changes[:20]:  # Limit output
                print(f"  [{change['col_letter']}{change['row']}] {change['label']}: "
                      f"'{change['old_value']}' -> '{change['new_value']}'")
            if len(report.is_changes) > 20:
                print(f"  ... and {len(report.is_changes) - 20} more")

        if report.bs_changes:
            print(f"\nBalance Sheet value changes ({len(report.bs_changes)}):")
            for change in report.bs_changes[:20]:
                print(f"  [{change['col_letter']}{change['row']}] {change['label']}: "
                      f"'{change['old_value']}' -> '{change['new_value']}'")
            if len(report.bs_changes) > 20:
                print(f"  ... and {len(report.bs_changes) - 20} more")

    if args.output:
        with open(args.output, 'w') as f:
            json.dump(report.to_dict(), f, indent=2)
        logger.info(f"Diff report saved to: {args.output}")


def main():
    parser = argparse.ArgumentParser(
        description="Validate and map net-net analysis workbook structures"
    )
    subparsers = parser.add_subparsers(dest='command', help='Commands')

    # Map command
    map_parser = subparsers.add_parser('map', help='Map workbook structure to JSON')
    map_parser.add_argument('--workbook', '-w', required=True,
                            help='Path to existing workbook')
    map_parser.add_argument('--output', '-o',
                            help='Output JSON file (default: <workbook>_structure.json)')

    # Validate command
    validate_parser = subparsers.add_parser('validate',
                                            help='Validate export against structure')
    validate_parser.add_argument('--structure', '-s', required=True,
                                 help='Path to structure JSON file')
    validate_parser.add_argument('--is-file', help='Path to Income Statement export')
    validate_parser.add_argument('--bs-file', help='Path to Balance Sheet export')
    validate_parser.add_argument('--output', '-o', help='Output validation results to JSON')

    # Diff command
    diff_parser = subparsers.add_parser('diff', help='Generate diff report')
    diff_parser.add_argument('--workbook', '-w', required=True,
                             help='Path to existing workbook')
    diff_parser.add_argument('--is-file', help='Path to Income Statement export')
    diff_parser.add_argument('--bs-file', help='Path to Balance Sheet export')
    diff_parser.add_argument('--output', '-o', help='Output diff report to JSON')
    diff_parser.add_argument('--verbose', '-v', action='store_true',
                             help='Show detailed changes')

    args = parser.parse_args()

    if args.command == 'map':
        cmd_map(args)
    elif args.command == 'validate':
        return cmd_validate(args)
    elif args.command == 'diff':
        cmd_diff(args)
    else:
        parser.print_help()
        return 1

    return 0


if __name__ == '__main__':
    sys.exit(main())
