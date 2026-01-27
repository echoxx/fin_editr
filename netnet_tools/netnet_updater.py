#!/usr/bin/env python3
"""
netnet_updater.py - Safely updates net-net analysis workbooks with new data

This tool:
1. Validates structure compatibility before making changes
2. Creates a timestamped backup of the original workbook
3. Updates raw data cells in IS and BS sheets
4. Optionally extends formulas to new date columns
5. Outputs a detailed summary of all changes

Usage:
    # Update workbook with new data (validation only - dry run)
    python netnet_updater.py --workbook almedio.xlsx \
        --is-file "Almedio Inc - Income Statement.xlsx" \
        --bs-file "Almedio Inc - Balance Sheet.xlsx" \
        --dry-run

    # Actually perform the update
    python netnet_updater.py --workbook almedio.xlsx \
        --is-file "Almedio Inc - Income Statement.xlsx" \
        --bs-file "Almedio Inc - Balance Sheet.xlsx"

    # Include new periods (extends columns)
    python netnet_updater.py --workbook almedio.xlsx \
        --is-file "Almedio Inc - Income Statement.xlsx" \
        --bs-file "Almedio Inc - Balance Sheet.xlsx" \
        --extend-periods

    # Replace all data (when switching from annual to quarterly, etc.)
    python netnet_updater.py --workbook almedio.xlsx \
        --is-file "Almedio Inc - Income Statement.xlsx" \
        --bs-file "Almedio Inc - Balance Sheet.xlsx" \
        --replace-all
"""

import argparse
import json
import logging
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Import from validator
from netnet_validator import (
    WorkbookStructure,
    map_workbook_structure,
    validate_export_against_structure,
    generate_diff_report,
    find_sheets,
    extract_row_labels,
    extract_column_dates,
    extract_period_codes
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class UpdateResult:
    """Holds results from an update operation."""

    def __init__(self):
        self.success = False
        self.backup_path: str | None = None
        self.cells_updated = 0
        self.is_cells_updated = 0
        self.bs_cells_updated = 0
        self.new_columns_added = 0
        self.formulas_extended = 0
        self.sheets_renamed = 0
        self.formula_refs_updated = 0
        self.price_formulas_updated = 0
        self.old_company_prefix: str | None = None
        self.new_company_prefix: str | None = None
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.changes_log: list[dict] = []

    def to_dict(self) -> dict:
        return {
            "success": self.success,
            "backup_path": self.backup_path,
            "cells_updated": self.cells_updated,
            "is_cells_updated": self.is_cells_updated,
            "bs_cells_updated": self.bs_cells_updated,
            "new_columns_added": self.new_columns_added,
            "formulas_extended": self.formulas_extended,
            "sheets_renamed": self.sheets_renamed,
            "formula_refs_updated": self.formula_refs_updated,
            "price_formulas_updated": self.price_formulas_updated,
            "old_company_prefix": self.old_company_prefix,
            "new_company_prefix": self.new_company_prefix,
            "errors": self.errors,
            "warnings": self.warnings,
            "changes_log": self.changes_log
        }


def create_backup(workbook_path: str) -> str:
    """Create a timestamped backup of the workbook."""
    path = Path(workbook_path)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"{path.stem}_backup_{timestamp}{path.suffix}"
    backup_path = path.parent / backup_name

    shutil.copy2(workbook_path, backup_path)
    logger.info(f"Backup created: {backup_path}")

    return str(backup_path)


def is_formula(value) -> bool:
    """Check if a cell value is a formula."""
    return isinstance(value, str) and value.startswith('=')


def get_last_data_column(worksheet, date_row: int = 10) -> int:
    """Find the last column with date data."""
    last_col = 3  # Column C is the minimum
    for col in range(4, 100):  # Check up to column CV
        if worksheet.cell(row=date_row, column=col).value:
            last_col = col
        else:
            # Allow for one empty column but stop at second
            next_val = worksheet.cell(row=date_row, column=col + 1).value
            if not next_val:
                break
    return last_col


def copy_column_formatting(worksheet, source_col: int, target_col: int, max_row: int = 60):
    """Copy column formatting from source to target column."""
    for row in range(1, max_row + 1):
        source_cell = worksheet.cell(row=row, column=source_col)
        target_cell = worksheet.cell(row=row, column=target_col)

        # Copy number format
        target_cell.number_format = source_cell.number_format

        # Copy font (create a new font object to avoid shared reference issues)
        if source_cell.font:
            target_cell.font = openpyxl.styles.Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )

        # Copy fill
        if source_cell.fill and source_cell.fill.fill_type:
            target_cell.fill = openpyxl.styles.PatternFill(
                fill_type=source_cell.fill.fill_type,
                fgColor=source_cell.fill.fgColor,
                bgColor=source_cell.fill.bgColor
            )

        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = openpyxl.styles.Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )


def adjust_formula_column(formula: str, source_col: int, target_col: int) -> str:
    """Adjust a formula's column references from source to target column.

    Handles formulas like:
    - =_xlfn.NUMBERVALUE(almedio_bs!D12)
    - =D3*$C$3+$C$4*D4+D2
    - =AVERAGE(C7:D7)
    """
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)

    # Pattern to match column references (not absolute $X$ references)
    # We need to be careful not to replace column letters in the middle of words

    # First, handle sheet references like almedio_bs!D12
    def replace_sheet_ref(match):
        sheet = match.group(1)
        col = match.group(2)
        row = match.group(3)
        if col == source_letter:
            return f"{sheet}!{target_letter}{row}"
        return match.group(0)

    formula = re.sub(r'(\w+)!([A-Z]+)(\d+)', replace_sheet_ref, formula)

    # Handle range references like C7:D7
    def replace_range_ref(match):
        col1 = match.group(1)
        row1 = match.group(2)
        col2 = match.group(3)
        row2 = match.group(4)
        if col1 == source_letter:
            col1 = target_letter
        if col2 == source_letter:
            col2 = target_letter
        return f"{col1}{row1}:{col2}{row2}"

    formula = re.sub(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', replace_range_ref, formula)

    # Handle simple references like D3 (not absolute like $D$3)
    # Use word boundary to avoid replacing in the middle of words
    def replace_simple_ref(match):
        prefix = match.group(1)
        col = match.group(2)
        row = match.group(3)
        if col == source_letter:
            return f"{prefix}{target_letter}{row}"
        return match.group(0)

    # Match column references that aren't absolute (no $ prefix)
    # The prefix captures anything that's not a letter (including start of string)
    formula = re.sub(r'(^|[^A-Z$])([A-Z]+)(\d+)(?![A-Z])', replace_simple_ref, formula)

    return formula


def extend_formulas_to_column(
    worksheet,
    source_col: int,
    target_col: int,
    formula_rows: list[int]
) -> int:
    """Extend formulas from source column to target column."""
    formulas_extended = 0

    for row in formula_rows:
        source_cell = worksheet.cell(row=row, column=source_col)
        if is_formula(source_cell.value):
            new_formula = adjust_formula_column(source_cell.value, source_col, target_col)
            worksheet.cell(row=row, column=target_col).value = new_formula
            formulas_extended += 1
            logger.debug(f"Extended formula to {get_column_letter(target_col)}{row}: {new_formula}")

    return formulas_extended


def update_raw_data_sheet(
    existing_ws,
    export_ws,
    result: UpdateResult,
    sheet_type: str,
    dry_run: bool = False
) -> list[dict]:
    """Update raw data in a worksheet from export data.

    Returns list of changes made.
    """
    changes = []

    # Build date-to-column mappings
    existing_dates = extract_column_dates(existing_ws)
    export_dates = extract_column_dates(export_ws)

    date_to_existing_col = {date: col for col, date in existing_dates.items()}
    date_to_export_col = {date: col for col, date in export_dates.items()}

    # Get row labels from existing sheet
    row_labels = extract_row_labels(existing_ws)

    # Update values for matching dates
    for date, export_col in date_to_export_col.items():
        if date not in date_to_existing_col:
            continue  # Skip new dates for now

        existing_col = date_to_existing_col[date]

        for row, label in row_labels.items():
            existing_cell = existing_ws.cell(row=row, column=existing_col)
            export_cell = export_ws.cell(row=row, column=export_col)

            # Skip if existing cell has a formula (we never overwrite formulas)
            if is_formula(existing_cell.value):
                continue

            existing_val = existing_cell.value
            new_val = export_cell.value

            # Normalize for comparison
            existing_str = str(existing_val).strip() if existing_val else ""
            new_str = str(new_val).strip() if new_val else ""

            if existing_str != new_str:
                change = {
                    "sheet_type": sheet_type,
                    "row": row,
                    "col": existing_col,
                    "col_letter": get_column_letter(existing_col),
                    "date": date,
                    "label": label,
                    "old_value": existing_val,
                    "new_value": new_val
                }
                changes.append(change)

                if not dry_run:
                    existing_cell.value = new_val
                    logger.debug(f"Updated {get_column_letter(existing_col)}{row}: "
                                f"'{existing_val}' -> '{new_val}'")

    return changes


def normalize_label(label: str) -> str:
    """Normalize a label for matching (lowercase, strip, remove extra spaces)."""
    if not label:
        return ""
    return " ".join(label.lower().strip().split())


def extract_company_name_from_filename(filename: str) -> str | None:
    """Extract company name from an Investing.com export filename.

    Expected formats:
    - "Nansin Co Ltd - Balance Sheet.xlsx"
    - "Almedio Inc - Income Statement.xlsx"
    - "Car Mate Mfg Co Ltd - Balance Sheet.xlsx"

    Returns the company name portion, or None if not parseable.
    """
    if not filename:
        return None

    # Get just the filename without path
    basename = Path(filename).stem  # e.g., "Nansin Co Ltd - Balance Sheet"

    # Split on " - " to separate company name from statement type
    parts = basename.split(" - ")
    if len(parts) >= 2:
        return parts[0].strip()

    return None


def normalize_company_name_for_sheet(company_name: str) -> str:
    """Convert company name to a valid sheet name prefix.

    Examples:
    - "Nansin Co Ltd" -> "nansin"
    - "Almedio Inc" -> "almedio"
    - "Car Mate Mfg Co Ltd" -> "carmate"
    """
    if not company_name:
        return ""

    # Take first word and lowercase it
    # For multi-word companies like "Car Mate", concatenate them
    words = company_name.lower().split()

    # Common suffixes to remove
    suffixes_to_remove = {'co', 'ltd', 'inc', 'corp', 'mfg', 'corporation', 'company', 'limited'}

    # Filter out suffixes and take remaining words
    name_words = [w for w in words if w not in suffixes_to_remove]

    if not name_words:
        # If all words were suffixes, just use the first word
        name_words = [words[0]] if words else ['unknown']

    # Join remaining words (handles "Car Mate" -> "carmate")
    return ''.join(name_words)


def find_sheet_name_references(workbook: openpyxl.Workbook, old_prefix: str) -> list[tuple[str, int, int, str]]:
    """Find all formula references to sheets with the given prefix.

    Returns list of (sheet_name, row, col, formula) tuples.
    """
    references = []
    old_is_pattern = f"{old_prefix}_IS" if old_prefix else None
    old_bs_pattern = f"{old_prefix}_bs" if old_prefix else None

    # Also check case variations
    patterns = set()
    if old_prefix:
        patterns.add(f"{old_prefix}_IS")
        patterns.add(f"{old_prefix}_is")
        patterns.add(f"{old_prefix}_bs")
        patterns.add(f"{old_prefix}_BS")
        patterns.add(f"{old_prefix.lower()}_IS")
        patterns.add(f"{old_prefix.lower()}_is")
        patterns.add(f"{old_prefix.lower()}_bs")
        patterns.add(f"{old_prefix.lower()}_BS")

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in range(1, 100):
            for col in range(1, 100):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    # Check if formula contains any of the old sheet name patterns
                    for pattern in patterns:
                        if pattern in formula:
                            references.append((sheet_name, row, col, formula))
                            break

    return references


def rename_sheets_and_update_references(
    workbook: openpyxl.Workbook,
    old_prefix: str,
    new_prefix: str,
    dry_run: bool = False
) -> tuple[int, int]:
    """Rename raw data sheets and update all formula references.

    Returns (sheets_renamed, formulas_updated) counts.
    """
    sheets_renamed = 0
    formulas_updated = 0

    # Build mapping of old sheet names to new sheet names
    rename_map = {}
    for sheet_name in workbook.sheetnames:
        sheet_lower = sheet_name.lower()
        # Check for IS sheets
        if sheet_lower == f"{old_prefix.lower()}_is":
            rename_map[sheet_name] = f"{new_prefix}_IS"
        # Check for BS sheets
        elif sheet_lower == f"{old_prefix.lower()}_bs":
            rename_map[sheet_name] = f"{new_prefix}_bs"

    if not rename_map:
        logger.warning(f"No sheets found matching prefix '{old_prefix}'")
        return 0, 0

    logger.info(f"Sheet rename map: {rename_map}")

    # Update all formula references first
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in range(1, 100):
            for col in range(1, 100):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    original_formula = cell.value
                    new_formula = original_formula

                    # Replace all old sheet name references with new ones
                    for old_name, new_name in rename_map.items():
                        # Handle various cases: exact match and case variations
                        new_formula = new_formula.replace(f"{old_name}!", f"{new_name}!")
                        # Also try lowercase variations
                        new_formula = new_formula.replace(f"{old_name.lower()}!", f"{new_name}!")

                    if new_formula != original_formula:
                        if not dry_run:
                            cell.value = new_formula
                        formulas_updated += 1
                        logger.debug(f"Updated formula in {sheet_name}!{get_column_letter(col)}{row}")

    # Now rename the sheets
    for old_name, new_name in rename_map.items():
        if old_name in workbook.sheetnames:
            if not dry_run:
                workbook[old_name].title = new_name
            sheets_renamed += 1
            logger.info(f"Renamed sheet '{old_name}' to '{new_name}'")

    return sheets_renamed, formulas_updated


def detect_current_company_prefix(workbook: openpyxl.Workbook) -> str | None:
    """Detect the current company prefix from sheet names.

    Looks for sheets matching patterns like 'xxx_IS' or 'xxx_bs'.
    """
    for sheet_name in workbook.sheetnames:
        sheet_lower = sheet_name.lower()
        if sheet_lower.endswith('_is'):
            return sheet_name[:-3]  # Remove '_is' suffix, preserve original case
        elif sheet_lower.endswith('_bs'):
            return sheet_name[:-3]  # Remove '_bs' suffix
    return None


def build_label_to_row_map(worksheet, start_row: int = 12, max_row: int = 60, label_col: int = 3) -> dict[str, int]:
    """Build a mapping of normalized label -> row number."""
    label_map = {}
    for row in range(start_row, max_row + 1):
        value = worksheet.cell(row=row, column=label_col).value
        if value and isinstance(value, str) and value.strip():
            normalized = normalize_label(value)
            if normalized not in label_map:  # Keep first occurrence
                label_map[normalized] = row
    return label_map


def replace_all_raw_data(
    existing_ws,
    export_ws,
    result: UpdateResult,
    sheet_type: str,
    dry_run: bool = False
) -> int:
    """Replace all raw data in a worksheet with export data.

    Matches data by row label (not row number) to handle different export structures.
    Clears existing data and copies all data from export.
    Returns count of cells written.
    """
    cells_written = 0

    # Clear existing data (columns D onwards, rows 8, 10, and 12-60)
    if not dry_run:
        for col in range(4, 100):
            existing_ws.cell(row=8, column=col).value = None   # period codes
            existing_ws.cell(row=10, column=col).value = None  # dates
        for row in range(12, 60):
            for col in range(4, 100):
                cell = existing_ws.cell(row=row, column=col)
                if not is_formula(cell.value):
                    cell.value = None

    # Copy period codes (row 8)
    for col in range(4, 100):
        val = export_ws.cell(row=8, column=col).value
        if val:
            if not dry_run:
                existing_ws.cell(row=8, column=col).value = val
            cells_written += 1
        elif col > 4:
            break

    # Copy dates (row 10)
    for col in range(4, 100):
        val = export_ws.cell(row=10, column=col).value
        if val:
            if not dry_run:
                existing_ws.cell(row=10, column=col).value = val
            cells_written += 1
        elif col > 4:
            break

    # Build label-to-row mappings for both worksheets
    existing_labels = build_label_to_row_map(existing_ws)
    export_labels = build_label_to_row_map(export_ws)

    # Build reverse map: export row -> existing row (matched by label)
    export_row_to_existing_row = {}
    for label, export_row in export_labels.items():
        if label in existing_labels:
            export_row_to_existing_row[export_row] = existing_labels[label]

    logger.debug(f"Label matching: {len(export_row_to_existing_row)} rows matched by label")

    # Copy data rows by matching labels
    export_dates = extract_column_dates(export_ws)
    max_col = max(export_dates.keys()) if export_dates else 4

    for export_row in range(12, 60):
        # Find the corresponding row in existing worksheet by label
        if export_row in export_row_to_existing_row:
            existing_row = export_row_to_existing_row[export_row]
        else:
            # No label match - skip this row (don't copy unmatched data)
            continue

        for col in range(4, max_col + 1):
            val = export_ws.cell(row=export_row, column=col).value
            if val is not None:
                if not dry_run:
                    existing_ws.cell(row=existing_row, column=col).value = val
                cells_written += 1

    logger.info(f"Replaced {sheet_type.upper()} data: {cells_written} cells (label-matched)")
    return cells_written


def sync_calculation_sheet_dates(
    workbook: openpyxl.Workbook,
    is_sheet_name: str | None,
    bs_sheet_name: str | None,
    calculation_sheets: list[str],
    dry_run: bool = False
) -> int:
    """Sync date headers in calculation sheets to match raw data sheets.

    Returns count of date headers updated.
    """
    # Get dates from raw data sheet (prefer IS, fall back to BS)
    source_sheet = is_sheet_name or bs_sheet_name
    if not source_sheet or source_sheet not in workbook.sheetnames:
        logger.warning("No source sheet found for date sync")
        return 0

    source_ws = workbook[source_sheet]
    dates = []
    for col in range(4, 100):
        val = source_ws.cell(row=10, column=col).value
        if val:
            if hasattr(val, 'strftime'):
                dates.append(val.strftime('%Y-%m-%d'))
            else:
                dates.append(str(val)[:10] if val else None)
        elif col > 4:
            break

    if not dates:
        logger.warning("No dates found in source sheet")
        return 0

    logger.info(f"Syncing {len(dates)} dates to calculation sheets")

    dates_updated = 0

    # Configuration for each calculation sheet
    # sheet_name: (date_row, start_col)
    calc_sheet_config = {
        'ncav': (1, 4),
        'profitability': (1, 3),
        'ro': (1, 3),
    }

    for sheet_name in calculation_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        if sheet_name not in calc_sheet_config:
            continue

        date_row, start_col = calc_sheet_config[sheet_name]
        ws = workbook[sheet_name]

        if not dry_run:
            # Clear old dates first
            for col in range(start_col, 100):
                ws.cell(row=date_row, column=col).value = None

            # Write new dates
            for i, date in enumerate(dates):
                ws.cell(row=date_row, column=start_col + i).value = date
                dates_updated += 1

        logger.info(f"  Updated {sheet_name} with {len(dates)} date headers")

    return dates_updated


def extend_calculation_sheet_formulas(
    workbook: openpyxl.Workbook,
    is_sheet_name: str | None,
    bs_sheet_name: str | None,
    calculation_sheets: list[str],
    dry_run: bool = False
) -> int:
    """Extend formulas in calculation sheets to cover all date columns.

    Returns count of formulas extended.
    """
    # Get number of periods from raw data
    source_sheet = is_sheet_name or bs_sheet_name
    if not source_sheet or source_sheet not in workbook.sheetnames:
        return 0

    source_ws = workbook[source_sheet]
    num_periods = 0
    for col in range(4, 100):
        if source_ws.cell(row=10, column=col).value:
            num_periods += 1
        elif col > 4:
            break

    if num_periods == 0:
        return 0

    logger.info(f"Extending formulas to cover {num_periods} periods")

    # Configuration: sheet_name -> (start_col, max_row)
    calc_config = {
        'ncav': (4, 30),
        'profitability': (3, 15),
        'ro': (3, 20),
    }

    total_formulas = 0

    for sheet_name in calculation_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        if sheet_name not in calc_config:
            continue

        start_col, max_row = calc_config[sheet_name]
        ws = workbook[sheet_name]
        target_col = start_col + num_periods - 1

        # Find last column with formulas
        last_formula_col = start_col
        for col in range(start_col, 100):
            has_formula = False
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    has_formula = True
                    break
            if has_formula:
                last_formula_col = col
            elif col > start_col + 5:
                break

        if last_formula_col >= target_col:
            continue

        if dry_run:
            logger.info(f"  Would extend {sheet_name} from col {get_column_letter(last_formula_col)} to {get_column_letter(target_col)}")
            continue

        # Extend formulas column by column
        formulas_added = 0
        for col in range(last_formula_col + 1, target_col + 1):
            for row in range(1, max_row + 1):
                source_cell = ws.cell(row=row, column=col - 1)
                target_cell = ws.cell(row=row, column=col)

                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    new_formula = adjust_formula_column(source_cell.value, col - 1, col)
                    target_cell.value = new_formula
                    formulas_added += 1

        logger.info(f"  Extended {sheet_name}: col {get_column_letter(last_formula_col)} -> {get_column_letter(target_col)} ({formulas_added} formulas)")
        total_formulas += formulas_added

    return total_formulas


def add_new_period_column(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    new_date: str,
    new_period_code: str,
    export_ws,
    export_col: int,
    calculation_sheets: list[str],
    result: UpdateResult,
    dry_run: bool = False
) -> int:
    """Add a new period column to the workbook.

    Returns number of formulas extended.
    """
    ws = workbook[sheet_name]

    # Find last data column
    last_col = get_last_data_column(ws)
    new_col = last_col + 1

    logger.info(f"Adding new column {get_column_letter(new_col)} for {new_date}")

    if dry_run:
        return 0

    # Copy column formatting
    copy_column_formatting(ws, last_col, new_col)

    # Add period code (row 8)
    ws.cell(row=8, column=new_col).value = new_period_code

    # Add date (row 10)
    ws.cell(row=10, column=new_col).value = new_date

    # Copy data from export
    row_labels = extract_row_labels(ws)
    for row in row_labels.keys():
        export_val = export_ws.cell(row=row, column=export_col).value
        ws.cell(row=row, column=new_col).value = export_val

    result.new_columns_added += 1

    # Extend formulas in calculation sheets
    formulas_extended = 0
    for calc_sheet_name in calculation_sheets:
        if calc_sheet_name not in workbook.sheetnames:
            continue

        calc_ws = workbook[calc_sheet_name]

        # Find the corresponding column in calculation sheet by date
        # First, find where dates are stored (usually row 1)
        calc_last_col = get_last_data_column(calc_ws, date_row=1)

        if calc_last_col < 4:
            continue

        # Check if calculation sheet has date headers that match raw data
        # Extend formula column if needed
        calc_new_col = calc_last_col + 1

        # Add date header
        calc_ws.cell(row=1, column=calc_new_col).value = new_date

        # Extend all formulas from last column
        for row in range(2, 60):
            source_cell = calc_ws.cell(row=row, column=calc_last_col)
            if is_formula(source_cell.value):
                new_formula = adjust_formula_column(source_cell.value, calc_last_col, calc_new_col)
                calc_ws.cell(row=row, column=calc_new_col).value = new_formula
                formulas_extended += 1

    return formulas_extended


def update_ncav_price_formulas(
    workbook: openpyxl.Workbook,
    dry_run: bool = False
) -> int:
    """Update price-related formulas in ncav sheet to reference the most recent period.

    The ncav sheet has special formulas that divide price by per-share values.
    These formulas should always reference the LAST column with data (most recent period),
    not follow the standard column-shifting pattern.

    Returns count of formulas updated.
    """
    if 'ncav' not in workbook.sheetnames:
        logger.debug("No ncav sheet found, skipping price formula update")
        return 0

    ws = workbook['ncav']

    # Find the last data column by looking at row 1 (date headers)
    last_data_col = 4  # Minimum is column D
    for col in range(4, 100):
        if ws.cell(row=1, column=col).value:
            last_data_col = col
        elif col > 5:
            # Allow one empty column but stop at second
            next_val = ws.cell(row=1, column=col + 1).value
            if not next_val:
                break

    last_col_letter = get_column_letter(last_data_col)
    logger.info(f"ncav sheet: last data column is {last_col_letter} (column {last_data_col})")

    # Find the Price row (look for "Price" label in column A or B)
    price_row = None
    for row in range(40, 50):
        label_a = ws.cell(row=row, column=1).value
        label_b = ws.cell(row=row, column=2).value
        label = (label_a or label_b or "").lower().strip()
        if label == "price":
            price_row = row
            break

    if not price_row:
        logger.warning("Could not find Price row in ncav sheet")
        return 0

    logger.debug(f"Found Price row at row {price_row}")

    # Define the rows with price-related formulas (P/NCAV, P/Discounted NCAV, P/NTA, P/Net Cash)
    # These are typically the rows immediately after Price
    price_formula_rows = []
    for row in range(price_row + 1, price_row + 10):
        label_a = ws.cell(row=row, column=1).value
        label_b = ws.cell(row=row, column=2).value
        label = (label_a or label_b or "").lower().strip()
        if label.startswith("p /") or label.startswith("p/"):
            price_formula_rows.append(row)

    if not price_formula_rows:
        logger.debug("No price ratio rows found")
        return 0

    logger.info(f"Found price ratio rows: {price_formula_rows}")

    # Define the mapping of ratio rows to their denominator rows
    # P/NCAV -> NCAV/Share (row 34)
    # P/Discounted NCAV -> Discounted NCAV/Share (row 38)
    # P/NTA -> Net Tangible Assets/Share (row 29)
    # P/Net Cash -> Net Cash/Share (row 41)
    ratio_to_denominator = {}
    for row in range(1, 60):
        label_a = ws.cell(row=row, column=1).value
        label_b = ws.cell(row=row, column=2).value
        label = (label_a or label_b or "").lower().strip()
        if "ncav / share" in label or "ncav/share" in label:
            if "discounted" in label:
                ratio_to_denominator["p / discounted ncav"] = row
                ratio_to_denominator["p/discounted ncav"] = row
            else:
                ratio_to_denominator["p / ncav"] = row
                ratio_to_denominator["p/ncav"] = row
        elif "net tangible assets / sha" in label or "nta / sha" in label or "nta/sha" in label:
            ratio_to_denominator["p / nta"] = row
            ratio_to_denominator["p/nta"] = row
        elif "net cash / share" in label or "net cash/share" in label:
            ratio_to_denominator["p / net cash"] = row
            ratio_to_denominator["p/net cash"] = row

    logger.debug(f"Ratio to denominator mapping: {ratio_to_denominator}")

    formulas_updated = 0

    # Update each price formula row
    for row in price_formula_rows:
        label_a = ws.cell(row=row, column=1).value
        label_b = ws.cell(row=row, column=2).value
        label = (label_a or label_b or "").lower().strip()

        # Find the denominator row for this ratio
        denom_row = ratio_to_denominator.get(label)

        if denom_row:
            # Build the new formula: =<price_col><price_row>/<price_col><denom_row>
            # Handle Net Cash specially with IFERROR
            if "net cash" in label:
                new_formula = f'=IFERROR({last_col_letter}{price_row}/{last_col_letter}{denom_row}, "NA")'
            else:
                new_formula = f"={last_col_letter}{price_row}/{last_col_letter}{denom_row}"

            current_formula = ws.cell(row=row, column=last_data_col).value

            if current_formula != new_formula:
                logger.info(f"Updating row {row} ({label}): {current_formula} -> {new_formula}")
                if not dry_run:
                    ws.cell(row=row, column=last_data_col).value = new_formula
                formulas_updated += 1
        else:
            logger.warning(f"Could not find denominator row for '{label}'")

    return formulas_updated


def update_workbook(
    workbook_path: str,
    is_export_path: str | None = None,
    bs_export_path: str | None = None,
    extend_periods: bool = False,
    replace_all: bool = False,
    dry_run: bool = False,
    force: bool = False,
    company_name: str | None = None
) -> UpdateResult:
    """Update a workbook with new data from Investing.com exports.

    Args:
        workbook_path: Path to the workbook to update
        is_export_path: Path to Income Statement export file
        bs_export_path: Path to Balance Sheet export file
        extend_periods: Add new period columns if found in export
        replace_all: Replace all data (for switching companies or data sources)
        dry_run: Show what would change without making changes
        force: Force update even if validation fails
        company_name: New company name (auto-detected from export filenames if not provided)
    """

    result = UpdateResult()

    # Step 1: Validate structure compatibility (skip if replace_all)
    if not replace_all:
        logger.info("Validating structure compatibility...")
        structure = map_workbook_structure(workbook_path)

        validation_errors = []
        if is_export_path:
            is_validation = validate_export_against_structure(structure, is_export_path, 'is')
            if not is_validation.is_valid and not force:
                validation_errors.extend(is_validation.errors)

        if bs_export_path:
            bs_validation = validate_export_against_structure(structure, bs_export_path, 'bs')
            if not bs_validation.is_valid and not force:
                validation_errors.extend(bs_validation.errors)

        if validation_errors and not force:
            result.errors = validation_errors
            result.errors.append("Validation failed. Use --force to override.")
            logger.error("Validation failed")
            return result
    else:
        logger.info("Replace-all mode: skipping validation")

    # Step 2: Create backup (unless dry run)
    if not dry_run:
        result.backup_path = create_backup(workbook_path)
    else:
        logger.info("DRY RUN - no backup created")

    # Step 3: Load workbooks
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    is_sheet, bs_sheet = find_sheets(wb)

    calculation_sheets = ['ncav', 'profitability', 'piotrosky', 'C7', 'ro']

    # Step 3.5: Handle company name / sheet renaming
    # Detect current company prefix from existing sheet names
    old_prefix = detect_current_company_prefix(wb)
    result.old_company_prefix = old_prefix

    # Determine new company name (from argument or auto-detect from export filenames)
    new_company_name = company_name
    if not new_company_name:
        # Try to extract from export filenames
        for export_path in [is_export_path, bs_export_path]:
            if export_path:
                extracted = extract_company_name_from_filename(export_path)
                if extracted:
                    new_company_name = extracted
                    break

    if new_company_name:
        new_prefix = normalize_company_name_for_sheet(new_company_name)
        result.new_company_prefix = new_prefix

        # Only rename if the prefix has changed
        if old_prefix and new_prefix and old_prefix.lower() != new_prefix.lower():
            logger.info(f"Company change detected: '{old_prefix}' -> '{new_prefix}'")
            sheets_renamed, formulas_updated = rename_sheets_and_update_references(
                wb, old_prefix, new_prefix, dry_run
            )
            result.sheets_renamed = sheets_renamed
            result.formula_refs_updated = formulas_updated

            # Update the sheet name variables to reflect the new names
            if sheets_renamed > 0:
                is_sheet, bs_sheet = find_sheets(wb)
                logger.info(f"Renamed {sheets_renamed} sheets, updated {formulas_updated} formula references")
        elif old_prefix and new_prefix:
            logger.info(f"Company prefix unchanged: '{old_prefix}'")
    else:
        logger.debug("No company name provided or detected from filenames")

    # Step 4: Update Income Statement
    if is_export_path and is_sheet:
        export_wb = openpyxl.load_workbook(is_export_path, data_only=False)

        if replace_all:
            logger.info("Replacing all Income Statement data...")
            cells = replace_all_raw_data(
                wb[is_sheet],
                export_wb.active,
                result,
                'is',
                dry_run
            )
            result.is_cells_updated = cells
        else:
            logger.info("Updating Income Statement data...")
            changes = update_raw_data_sheet(
                wb[is_sheet],
                export_wb.active,
                result,
                'is',
                dry_run
            )
            result.is_cells_updated = len(changes)
            result.changes_log.extend(changes)

            # Handle new periods if requested
            if extend_periods:
                existing_dates = set(extract_column_dates(wb[is_sheet]).values())
                export_dates = extract_column_dates(export_wb.active)
                export_codes = extract_period_codes(export_wb.active)

                for col, date in export_dates.items():
                    if date not in existing_dates:
                        logger.info(f"New period found: {date}")
                        formulas = add_new_period_column(
                            wb, is_sheet, date,
                            export_codes.get(col, ""),
                            export_wb.active, col,
                            calculation_sheets,
                            result, dry_run
                        )
                        result.formulas_extended += formulas

        export_wb.close()

    # Step 5: Update Balance Sheet
    if bs_export_path and bs_sheet:
        export_wb = openpyxl.load_workbook(bs_export_path, data_only=False)

        if replace_all:
            logger.info("Replacing all Balance Sheet data...")
            cells = replace_all_raw_data(
                wb[bs_sheet],
                export_wb.active,
                result,
                'bs',
                dry_run
            )
            result.bs_cells_updated = cells
        else:
            logger.info("Updating Balance Sheet data...")
            changes = update_raw_data_sheet(
                wb[bs_sheet],
                export_wb.active,
                result,
                'bs',
                dry_run
            )
            result.bs_cells_updated = len(changes)
            result.changes_log.extend(changes)

            # Handle new periods if requested
            if extend_periods:
                existing_dates = set(extract_column_dates(wb[bs_sheet]).values())
                export_dates = extract_column_dates(export_wb.active)
                export_codes = extract_period_codes(export_wb.active)

                for col, date in export_dates.items():
                    if date not in existing_dates:
                        # Only add if not already added by IS
                        if date not in set(extract_column_dates(wb[bs_sheet]).values()):
                            logger.info(f"New period found: {date}")
                            formulas = add_new_period_column(
                                wb, bs_sheet, date,
                                export_codes.get(col, ""),
                                export_wb.active, col,
                                calculation_sheets,
                                result, dry_run
                            )
                            result.formulas_extended += formulas

        export_wb.close()

    # Step 6: Sync calculation sheet dates and extend formulas
    result.cells_updated = result.is_cells_updated + result.bs_cells_updated

    if result.cells_updated > 0:
        logger.info("Syncing calculation sheet dates...")
        dates_synced = sync_calculation_sheet_dates(
            wb, is_sheet, bs_sheet, calculation_sheets, dry_run
        )
        if dates_synced > 0:
            logger.info(f"Synced {dates_synced} date headers in calculation sheets")

        # Extend formulas to cover all periods (especially important for replace_all)
        logger.info("Extending calculation sheet formulas...")
        formulas_extended = extend_calculation_sheet_formulas(
            wb, is_sheet, bs_sheet, calculation_sheets, dry_run
        )
        if formulas_extended > 0:
            result.formulas_extended += formulas_extended
            logger.info(f"Extended {formulas_extended} formulas in calculation sheets")

        # Update price-related formulas in ncav sheet to reference most recent period
        logger.info("Updating ncav price formulas...")
        price_formulas_updated = update_ncav_price_formulas(wb, dry_run)
        if price_formulas_updated > 0:
            result.price_formulas_updated = price_formulas_updated
            logger.info(f"Updated {price_formulas_updated} price formulas in ncav sheet")

    # Step 7: Save workbook
    if not dry_run and (result.cells_updated > 0 or result.sheets_renamed > 0):
        wb.save(workbook_path)
        logger.info(f"Workbook saved: {workbook_path}")

    wb.close()
    result.success = True

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Safely update net-net analysis workbooks with new data"
    )

    parser.add_argument('--workbook', '-w', required=True,
                        help='Path to workbook to update')
    parser.add_argument('--is-file', help='Path to Income Statement export')
    parser.add_argument('--bs-file', help='Path to Balance Sheet export')
    parser.add_argument('--extend-periods', action='store_true',
                        help='Add new period columns if found in export')
    parser.add_argument('--replace-all', action='store_true',
                        help='Replace all data (use when switching from annual to quarterly)')
    parser.add_argument('--dry-run', '-n', action='store_true',
                        help='Show what would be changed without making changes')
    parser.add_argument('--force', '-f', action='store_true',
                        help='Force update even if validation fails')
    parser.add_argument('--output', '-o',
                        help='Output update report to JSON file')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Show detailed changes')
    parser.add_argument('--company-name', '-c',
                        help='New company name (auto-detected from export filenames if not provided)')

    args = parser.parse_args()

    if not args.is_file and not args.bs_file:
        parser.error("At least one of --is-file or --bs-file is required")

    # Run the update
    result = update_workbook(
        args.workbook,
        args.is_file,
        args.bs_file,
        args.extend_periods,
        args.replace_all,
        args.dry_run,
        args.force,
        args.company_name
    )

    # Print summary
    print("\n" + "=" * 60)
    if args.dry_run:
        print("DRY RUN - No changes were made")
    else:
        print("UPDATE COMPLETE" if result.success else "UPDATE FAILED")
    print("=" * 60)

    if result.backup_path:
        print(f"Backup: {result.backup_path}")

    # Company name change info
    if result.sheets_renamed > 0:
        print(f"\nCompany Change:")
        print(f"  Old prefix: {result.old_company_prefix}")
        print(f"  New prefix: {result.new_company_prefix}")
        print(f"  Sheets renamed: {result.sheets_renamed}")
        print(f"  Formula references updated: {result.formula_refs_updated}")

    print(f"\nData Changes Summary:")
    print(f"  Income Statement cells updated: {result.is_cells_updated}")
    print(f"  Balance Sheet cells updated: {result.bs_cells_updated}")
    print(f"  Total cells updated: {result.cells_updated}")

    if args.extend_periods or result.formulas_extended > 0:
        print(f"  New columns added: {result.new_columns_added}")
        print(f"  Formulas extended: {result.formulas_extended}")

    if result.price_formulas_updated > 0:
        print(f"  Price formulas updated: {result.price_formulas_updated}")

    if result.errors:
        print(f"\nErrors:")
        for error in result.errors:
            print(f"  - {error}")

    if result.warnings:
        print(f"\nWarnings:")
        for warning in result.warnings:
            print(f"  - {warning}")

    if args.verbose and result.changes_log:
        print(f"\nDetailed Changes ({len(result.changes_log)}):")
        for change in result.changes_log[:50]:
            sheet = "IS" if change['sheet_type'] == 'is' else "BS"
            print(f"  [{sheet}] {change['col_letter']}{change['row']} "
                  f"({change['label']}): '{change['old_value']}' -> '{change['new_value']}'")
        if len(result.changes_log) > 50:
            print(f"  ... and {len(result.changes_log) - 50} more changes")

    if args.output:
        with open(args.output, 'w') as f:
            json.dump(result.to_dict(), f, indent=2, default=str)
        print(f"\nReport saved to: {args.output}")

    return 0 if result.success else 1


if __name__ == '__main__':
    sys.exit(main())
