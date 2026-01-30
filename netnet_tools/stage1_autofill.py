#!/usr/bin/env python3
"""
Stage 1 Quantitative Autofill Script

Automatically populates the Stage1_Quantitative sheet in the net-net diligence
checklist using data extracted from a company's analysis workbook (output from
the excel-financial-updater workflow).

Usage:
    python stage1_autofill.py toso/toso.xlsx --price 650 --market-cap 36
    python stage1_autofill.py toso/toso.xlsx --price 650 --market-cap 36 --dry-run
    python stage1_autofill.py toso/toso.xlsx --price 650 --market-cap 36 --report-only
"""

import argparse
import logging
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

# Import from sibling modules
from overview_populator import EXCHANGE_INFO

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# Data location configuration for company workbooks
DATA_LOCATIONS = {
    'ncav': {
        'shares_outstanding': 20,
        'net_current_assets': 23,
        'current_ratio': 26,
        'ncav_per_share': 34,
        'ncav_qoq_change': 35,
        'ncav_yoy_change': 36,
        'debt_to_assets': 52,
        'debt_to_equity': 53,
        'liabilities_to_equity': 54,
    },
    'profitability': {
        'net_income': 10,
    },
    'ro': {
        'roa_annual': 21,
    },
    'piotrosky': {
        'total_row': 10,
        'score_column': 11,  # Column K
    },
}

# Stage1_Quantitative cell locations
STAGE1_CELLS = {
    # Header section
    'company_name': 'B3',
    'ticker': 'B4',
    'date': 'B5',
    'exchange': 'E3',
    'currency': 'E4',
    'market_cap': 'E5',
    # Hard filters (values go in column C)
    'p_ncav': 'C12',
    'ncav_burn_rate': 'C13',
    'debt_equity': 'C14',
    'current_ratio': 'C15',
    # Soft filters (values go in column C)
    'roa_trend': 'C19',
    'piotroski_score': 'C20',
    'prior_price_above_ncav': 'C21',
    'positive_ttm_ni': 'C22',
    # NCAV Trajectory
    'ncav_current': 'B25',
    'ncav_1yr_ago': 'B26',
    'ncav_2yr_ago': 'B27',
}


@dataclass
class ExtractedData:
    """Container for all extracted data from company workbook."""
    # Header
    company_name: Optional[str] = None
    ticker: Optional[str] = None
    country: Optional[str] = None
    currency: Optional[str] = None
    exchange: Optional[str] = None

    # Hard filters
    ncav_per_share: Optional[float] = None
    p_ncav: Optional[float] = None
    ncav_burn_rate: Optional[float] = None
    debt_equity: Optional[float] = None
    current_ratio: Optional[float] = None

    # Soft filters
    roa_trend: Optional[str] = None
    piotroski_score: Optional[int] = None
    positive_ttm_ni: Optional[str] = None

    # NCAV trajectory
    ncav_current: Optional[float] = None
    ncav_1yr_ago: Optional[float] = None
    ncav_2yr_ago: Optional[float] = None

    # ROA values for trend display
    roa_values: Optional[list] = None


class CompanyWorkbookReader:
    """Reads data from company analysis workbook."""

    def __init__(self, workbook_path: str, price: float):
        self.path = Path(workbook_path)
        self.price = price

        # Load workbook with calculated values
        self.wb_values = openpyxl.load_workbook(str(self.path), data_only=True)

        # Find IS sheet name (ends with _IS or _is)
        self.is_sheet = None
        for name in self.wb_values.sheetnames:
            if name.lower().endswith('_is'):
                self.is_sheet = name
                break

        # Cache for data frequency detection
        self._data_frequency = {}

    def _parse_value(self, val) -> Optional[float]:
        """Parse a cell value to float, handling strings and percentages."""
        if val is None or val == '' or val == 'NA' or val == '-':
            return None
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            cleaned = val.replace('%', '').replace(',', '').strip()
            try:
                result = float(cleaned)
                if '%' in val:
                    result = result / 100
                return result
            except ValueError:
                return None
        return None

    def _get_cell_value(self, sheet_name: str, row: int, col: int) -> Optional[float]:
        """Get parsed value from a cell."""
        if sheet_name not in self.wb_values.sheetnames:
            return None
        ws = self.wb_values[sheet_name]
        val = ws.cell(row=row, column=col).value
        return self._parse_value(val)

    def _get_row_series(self, sheet_name: str, row_num: int,
                        start_col: int = 4, max_cols: int = 45) -> list:
        """
        Extract values from a row, handling None/empty cells.
        Returns list of (column_index, value) tuples for non-None values.
        """
        if sheet_name not in self.wb_values.sheetnames:
            return []

        ws = self.wb_values[sheet_name]
        series = []
        consecutive_empty = 0

        for col in range(start_col, start_col + max_cols):
            val = self._parse_value(ws.cell(row=row_num, column=col).value)
            if val is not None:
                series.append((col, val))
                consecutive_empty = 0
            else:
                consecutive_empty += 1
                if consecutive_empty >= 3 and series:
                    break

        return series

    def _detect_data_frequency(self, sheet_name: str, sample_row: int) -> str:
        """
        Detect if data is quarterly or semi-annual.
        Semi-annual data has pairs of identical consecutive values.
        """
        if sheet_name in self._data_frequency:
            return self._data_frequency[sheet_name]

        series = self._get_row_series(sheet_name, sample_row)
        if len(series) < 4:
            self._data_frequency[sheet_name] = 'quarterly'
            return 'quarterly'

        # Check last 8 values for identical consecutive pairs
        recent = series[-8:] if len(series) >= 8 else series
        pairs_identical = 0
        pairs_different = 0

        for i in range(0, len(recent) - 1, 2):
            if i + 1 < len(recent):
                val1 = recent[i][1]
                val2 = recent[i + 1][1]
                if val1 is not None and val2 is not None:
                    if abs(val1) > 0.0001:
                        if abs(val1 - val2) / abs(val1) < 0.0001:
                            pairs_identical += 1
                        else:
                            pairs_different += 1
                    elif abs(val2) < 0.0001:
                        pairs_identical += 1
                    else:
                        pairs_different += 1

        frequency = 'semi-annual' if pairs_identical > pairs_different else 'quarterly'
        self._data_frequency[sheet_name] = frequency
        logger.debug(f"Detected {frequency} data for {sheet_name}")
        return frequency

    def _get_yoy_step(self, sheet_name: str, sample_row: int) -> int:
        """Get column step for year-over-year comparison."""
        frequency = self._detect_data_frequency(sheet_name, sample_row)
        # For semi-annual: 2 periods = 1 year
        # For quarterly: 4 periods = 1 year
        return 2 if frequency == 'semi-annual' else 4

    def get_company_name(self) -> Optional[str]:
        """Extract company name from IS sheet C2."""
        if self.is_sheet and self.is_sheet in self.wb_values.sheetnames:
            ws = self.wb_values[self.is_sheet]
            return ws['C2'].value
        return None

    def get_overview_field(self, field: str) -> Optional[str]:
        """Read field from Overview sheet."""
        if 'Overview' not in self.wb_values.sheetnames:
            return None
        ws = self.wb_values['Overview']
        cell_map = {
            'ticker': 'C3',
            'country': 'C4',
            'currency': 'C5',
        }
        if field in cell_map:
            return ws[cell_map[field]].value
        return None

    def infer_exchange_from_country(self, country: str) -> Optional[str]:
        """Infer exchange code from country name."""
        country_to_exchange = {
            'Japan': 'TYO',
            'Hong Kong': 'HKG',
            'United States': 'NYSE',
            'United Kingdom': 'LON',
            'France': 'EPA',
            'Germany': 'ETR',
            'South Korea': 'KRX',
            'Singapore': 'SGX',
            'Taiwan': 'TPE',
            'Australia': 'ASX',
            'Canada': 'TSX',
        }
        return country_to_exchange.get(country)

    def get_latest_value(self, sheet_name: str, row: int) -> Optional[float]:
        """Get the most recent value from a row."""
        series = self._get_row_series(sheet_name, row)
        if series:
            return series[-1][1]
        return None

    def get_ncav_per_share(self) -> Optional[float]:
        """Get current NCAV per share from ncav sheet row 34."""
        return self.get_latest_value('ncav', DATA_LOCATIONS['ncav']['ncav_per_share'])

    def get_p_ncav(self) -> Optional[float]:
        """Calculate P/NCAV using provided price."""
        ncav = self.get_ncav_per_share()
        if ncav and ncav > 0:
            return self.price / ncav
        return None

    def get_ncav_burn_rate(self) -> Optional[float]:
        """Get NCAV YoY change from ncav sheet row 36."""
        # First try pre-calculated YoY change
        yoy = self.get_latest_value('ncav', DATA_LOCATIONS['ncav']['ncav_yoy_change'])
        if yoy is not None:
            return yoy

        # Fallback: calculate from NCAV series
        ncav_row = DATA_LOCATIONS['ncav']['ncav_per_share']
        series = self._get_row_series('ncav', ncav_row)
        yoy_step = self._get_yoy_step('ncav', ncav_row)

        if len(series) > yoy_step:
            current = series[-1][1]
            prior = series[-(yoy_step + 1)][1]
            if prior and prior != 0:
                return (current - prior) / abs(prior)
        return None

    def get_debt_equity(self) -> Optional[float]:
        """Get D/E ratio from ncav sheet row 53."""
        return self.get_latest_value('ncav', DATA_LOCATIONS['ncav']['debt_to_equity'])

    def get_current_ratio(self) -> Optional[float]:
        """Get current ratio from ncav sheet row 26."""
        return self.get_latest_value('ncav', DATA_LOCATIONS['ncav']['current_ratio'])

    def get_roa_trend(self) -> tuple[Optional[str], list]:
        """
        Analyze ROA trend over 3 years.
        Returns (trend_string, [roa_values])
        """
        roa_row = DATA_LOCATIONS['ro']['roa_annual']
        series = self._get_row_series('ro', roa_row)
        yoy_step = self._get_yoy_step('ro', roa_row)

        # Need at least 3 years of data
        min_periods = yoy_step * 2 + 1
        if len(series) < min_periods:
            return None, []

        # Get values: current, 1yr ago, 2yr ago
        current = series[-1][1]
        one_yr = series[-(yoy_step + 1)][1] if len(series) > yoy_step else None
        two_yr = series[-(yoy_step * 2 + 1)][1] if len(series) > yoy_step * 2 else None

        roa_values = [two_yr, one_yr, current]

        if current is None or one_yr is None or two_yr is None:
            return None, roa_values

        # Determine trend
        if current > one_yr and one_yr > two_yr:
            return "increasing", roa_values
        elif current < one_yr and one_yr < two_yr:
            return "declining", roa_values
        else:
            return "stable", roa_values

    def get_piotroski_score(self) -> Optional[int]:
        """Get Piotroski F-Score total from piotrosky sheet."""
        if 'piotrosky' not in self.wb_values.sheetnames:
            return None
        ws = self.wb_values['piotrosky']
        total_row = DATA_LOCATIONS['piotrosky']['total_row']
        score_col = DATA_LOCATIONS['piotrosky']['score_column']
        val = ws.cell(row=total_row, column=score_col).value
        if val is not None:
            try:
                return int(val)
            except (ValueError, TypeError):
                return None
        return None

    def has_positive_ttm_net_income(self) -> Optional[str]:
        """Check if TTM net income is positive."""
        if not self.is_sheet:
            return None

        # Net income is typically row 33 in IS sheet (Net Income to Stockholders)
        # Try profitability sheet first
        ni = self.get_latest_value('profitability', DATA_LOCATIONS['profitability']['net_income'])

        if ni is not None:
            return "Yes" if ni > 0 else "No"
        return None

    def get_ncav_trajectory(self) -> tuple[Optional[float], Optional[float], Optional[float]]:
        """
        Get NCAV values: current, 1 year ago, 2 years ago.
        Returns (current, 1yr_ago, 2yr_ago)
        """
        ncav_row = DATA_LOCATIONS['ncav']['ncav_per_share']
        series = self._get_row_series('ncav', ncav_row)
        yoy_step = self._get_yoy_step('ncav', ncav_row)

        current = series[-1][1] if series else None
        one_yr = series[-(yoy_step + 1)][1] if len(series) > yoy_step else None
        two_yr = series[-(yoy_step * 2 + 1)][1] if len(series) > yoy_step * 2 else None

        return current, one_yr, two_yr

    def extract_all(self) -> ExtractedData:
        """Extract all data needed for Stage1_Quantitative."""
        data = ExtractedData()

        # Header info
        data.company_name = self.get_company_name()
        data.ticker = self.get_overview_field('ticker')
        data.country = self.get_overview_field('country')
        data.currency = self.get_overview_field('currency')

        # Infer exchange from country
        if data.country:
            data.exchange = self.infer_exchange_from_country(data.country)

        # Hard filters
        data.ncav_per_share = self.get_ncav_per_share()
        data.p_ncav = self.get_p_ncav()
        data.ncav_burn_rate = self.get_ncav_burn_rate()
        data.debt_equity = self.get_debt_equity()
        data.current_ratio = self.get_current_ratio()

        # Soft filters
        data.roa_trend, data.roa_values = self.get_roa_trend()
        data.piotroski_score = self.get_piotroski_score()
        data.positive_ttm_ni = self.has_positive_ttm_net_income()

        # NCAV trajectory
        data.ncav_current, data.ncav_1yr_ago, data.ncav_2yr_ago = self.get_ncav_trajectory()

        return data

    def close(self):
        """Close the workbook."""
        self.wb_values.close()


class Stage1Writer:
    """Writes extracted data to Stage1_Quantitative sheet."""

    def __init__(self, checklist_path: str):
        self.path = Path(checklist_path)
        self.wb = openpyxl.load_workbook(str(self.path))

        if 'Stage1_Quantitative' not in self.wb.sheetnames:
            raise ValueError(f"Checklist has no 'Stage1_Quantitative' sheet: {self.path}")

        self.ws = self.wb['Stage1_Quantitative']

    def write_value(self, cell_ref: str, value) -> bool:
        """Write a value to a cell. Returns True if written."""
        if value is not None:
            self.ws[cell_ref] = value
            return True
        return False

    def populate(self, data: ExtractedData, market_cap: float) -> dict:
        """
        Populate Stage1_Quantitative sheet with extracted data.
        Returns dict of fields populated.
        """
        populated = {}

        # Header section
        if self.write_value(STAGE1_CELLS['company_name'], data.company_name):
            populated['company_name'] = data.company_name

        if self.write_value(STAGE1_CELLS['ticker'], data.ticker):
            populated['ticker'] = data.ticker

        # Set current date
        today = datetime.now().strftime('%Y-%m-%d')
        self.write_value(STAGE1_CELLS['date'], today)
        populated['date'] = today

        if self.write_value(STAGE1_CELLS['exchange'], data.exchange):
            populated['exchange'] = data.exchange

        if self.write_value(STAGE1_CELLS['currency'], data.currency):
            populated['currency'] = data.currency

        # Market cap (required parameter)
        self.write_value(STAGE1_CELLS['market_cap'], market_cap)
        populated['market_cap'] = market_cap

        # Hard filters
        if data.p_ncav is not None:
            self.write_value(STAGE1_CELLS['p_ncav'], round(data.p_ncav, 2))
            populated['p_ncav'] = round(data.p_ncav, 2)

        if data.ncav_burn_rate is not None:
            self.write_value(STAGE1_CELLS['ncav_burn_rate'], round(data.ncav_burn_rate, 4))
            populated['ncav_burn_rate'] = round(data.ncav_burn_rate, 4)

        if data.debt_equity is not None:
            self.write_value(STAGE1_CELLS['debt_equity'], round(data.debt_equity, 3))
            populated['debt_equity'] = round(data.debt_equity, 3)

        if data.current_ratio is not None:
            self.write_value(STAGE1_CELLS['current_ratio'], round(data.current_ratio, 2))
            populated['current_ratio'] = round(data.current_ratio, 2)

        # Soft filters
        if self.write_value(STAGE1_CELLS['roa_trend'], data.roa_trend):
            populated['roa_trend'] = data.roa_trend

        if data.piotroski_score is not None:
            self.write_value(STAGE1_CELLS['piotroski_score'], data.piotroski_score)
            populated['piotroski_score'] = data.piotroski_score

        # Leave prior_price_above_ncav blank for manual input

        if self.write_value(STAGE1_CELLS['positive_ttm_ni'], data.positive_ttm_ni):
            populated['positive_ttm_ni'] = data.positive_ttm_ni

        # NCAV trajectory
        if data.ncav_current is not None:
            self.write_value(STAGE1_CELLS['ncav_current'], round(data.ncav_current, 2))
            populated['ncav_current'] = round(data.ncav_current, 2)

        if data.ncav_1yr_ago is not None:
            self.write_value(STAGE1_CELLS['ncav_1yr_ago'], round(data.ncav_1yr_ago, 2))
            populated['ncav_1yr_ago'] = round(data.ncav_1yr_ago, 2)

        if data.ncav_2yr_ago is not None:
            self.write_value(STAGE1_CELLS['ncav_2yr_ago'], round(data.ncav_2yr_ago, 2))
            populated['ncav_2yr_ago'] = round(data.ncav_2yr_ago, 2)

        return populated

    def save(self, output_path: Optional[str] = None):
        """Save the workbook."""
        save_path = output_path or str(self.path)
        self.wb.save(save_path)
        logger.info(f"Saved checklist to: {save_path}")

    def close(self):
        """Close the workbook."""
        self.wb.close()


def evaluate_pass_fail(data: ExtractedData) -> dict:
    """Evaluate PASS/FAIL for hard and soft filters."""
    results = {}

    # Hard filters
    if data.p_ncav is not None:
        results['p_ncav'] = 'PASS' if data.p_ncav < 0.67 else 'FAIL'

    if data.ncav_burn_rate is not None:
        results['ncav_burn_rate'] = 'PASS' if data.ncav_burn_rate > -0.10 else 'FAIL'

    if data.debt_equity is not None:
        results['debt_equity'] = 'PASS' if data.debt_equity < 0.50 else 'FAIL'

    if data.current_ratio is not None:
        results['current_ratio'] = 'PASS' if data.current_ratio > 1.5 else 'FAIL'

    # Soft filter evaluation (for display)
    if data.piotroski_score is not None:
        results['piotroski'] = 'PASS' if data.piotroski_score >= 5 else 'FAIL'

    if data.positive_ttm_ni is not None:
        results['ttm_ni'] = 'PASS' if data.positive_ttm_ni == 'Yes' else 'FAIL'

    return results


def print_report(data: ExtractedData, market_cap: float, price: float,
                 populated: dict, output_path: str):
    """Print a formatted report of extracted data."""

    results = evaluate_pass_fail(data)

    print("\n" + "=" * 68)
    print(f"STAGE 1 QUANTITATIVE AUTOFILL - {data.company_name or 'Unknown'}")
    print("=" * 68)
    print(f"Input: {price} | Market Cap: {market_cap} USD mm")
    print()

    print("HARD FILTERS:")

    # P/NCAV
    if data.p_ncav is not None:
        pf = results.get('p_ncav', '?')
        print(f"  P/NCAV: {data.p_ncav:.2f} -> {pf}")
    else:
        print("  P/NCAV: [NOT AVAILABLE]")

    # NCAV Burn Rate
    if data.ncav_burn_rate is not None:
        pf = results.get('ncav_burn_rate', '?')
        print(f"  NCAV Burn Rate: {data.ncav_burn_rate:.2%} -> {pf}")
    else:
        print("  NCAV Burn Rate: [NOT AVAILABLE]")

    # D/E
    if data.debt_equity is not None:
        pf = results.get('debt_equity', '?')
        print(f"  Debt/Equity: {data.debt_equity:.2f} -> {pf}")
    else:
        print("  Debt/Equity: [NOT AVAILABLE]")

    # Current Ratio
    if data.current_ratio is not None:
        pf = results.get('current_ratio', '?')
        print(f"  Current Ratio: {data.current_ratio:.2f} -> {pf}")
    else:
        print("  Current Ratio: [NOT AVAILABLE]")

    print()
    print("SOFT FILTERS:")

    # ROA Trend
    if data.roa_trend:
        roa_str = ""
        if data.roa_values and all(v is not None for v in data.roa_values):
            roa_str = f" ({data.roa_values[0]:.1%} -> {data.roa_values[1]:.1%} -> {data.roa_values[2]:.1%})"
        print(f"  ROA Trend (3yr): {data.roa_trend}{roa_str}")
    else:
        print("  ROA Trend (3yr): [NOT AVAILABLE]")

    # Piotroski
    if data.piotroski_score is not None:
        print(f"  Piotroski F-Score: {data.piotroski_score}")
    else:
        print("  Piotroski F-Score: [NOT AVAILABLE]")

    # Prior Price > 1x NCAV
    print("  Prior Price > 1x NCAV: [MANUAL - left blank]")

    # TTM Net Income
    if data.positive_ttm_ni:
        print(f"  Positive TTM Net Income: {data.positive_ttm_ni}")
    else:
        print("  Positive TTM Net Income: [NOT AVAILABLE]")

    print()
    print("NCAV TRAJECTORY:")

    current_str = f"{data.ncav_current:.2f}" if data.ncav_current is not None else "N/A"
    one_yr_str = f"{data.ncav_1yr_ago:.2f}" if data.ncav_1yr_ago is not None else "N/A"
    two_yr_str = f"{data.ncav_2yr_ago:.2f}" if data.ncav_2yr_ago is not None else "N/A"

    print(f"  Current: {current_str} | 1yr ago: {one_yr_str} | 2yr ago: {two_yr_str}")

    # Calculate changes if we have the data
    if data.ncav_current is not None and data.ncav_1yr_ago is not None and data.ncav_1yr_ago != 0:
        yoy = (data.ncav_current - data.ncav_1yr_ago) / abs(data.ncav_1yr_ago)
        yoy_str = f"{yoy:.2%}"
    else:
        yoy_str = "N/A"

    if data.ncav_current is not None and data.ncav_2yr_ago is not None and data.ncav_2yr_ago > 0:
        cagr = (data.ncav_current / data.ncav_2yr_ago) ** 0.5 - 1
        cagr_str = f"{cagr:.2%}"
    else:
        cagr_str = "N/A"

    print(f"  YoY Change: {yoy_str} | 2-year CAGR: {cagr_str}")

    print()
    print(f"RESULT: {len(populated)} fields populated, 1 requires manual input")
    print(f"Output: {output_path}")
    print("=" * 68)


def create_backup(filepath: str) -> str:
    """Create a timestamped backup of a file."""
    path = Path(filepath)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"{path.stem}_backup_{timestamp}{path.suffix}"
    backup_path = path.parent / backup_name
    shutil.copy2(filepath, backup_path)
    return str(backup_path)


def validate_company_workbook(wb_path: str) -> list[str]:
    """Validate that company workbook has required sheets."""
    errors = []

    try:
        wb = openpyxl.load_workbook(wb_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return [f"Cannot open workbook: {e}"]

    required = ['ncav', 'ro', 'profitability', 'piotrosky', 'Overview']
    for sheet in required:
        if sheet not in sheets:
            errors.append(f"Missing required sheet: {sheet}")

    # Check for IS sheet
    has_is = any(name.lower().endswith('_is') for name in sheets)
    if not has_is:
        errors.append("Missing Income Statement sheet (ending in _IS)")

    return errors


def main():
    parser = argparse.ArgumentParser(
        description='Auto-populate Stage1_Quantitative sheet from company workbook'
    )
    parser.add_argument('workbook', help='Path to company analysis workbook')
    parser.add_argument('--price', '-p', type=float, required=True,
                        help='Current stock price')
    parser.add_argument('--market-cap', '-m', type=float, required=True,
                        help='Market cap in USD millions')
    parser.add_argument('--checklist', '-c',
                        help='Path to checklist template (default: ../netnet_diligence_checklist.xlsx)')
    parser.add_argument('--output', '-o',
                        help='Output path (default: [company]_netnet_diligence_checklist.xlsx in workbook folder)')
    parser.add_argument('--dry-run', '-n', action='store_true',
                        help='Show what would be written without modifying files')
    parser.add_argument('--report-only', '-r', action='store_true',
                        help='Extract and display without creating output file')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Show detailed debug output')

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Resolve paths
    workbook_path = Path(args.workbook).resolve()

    if not workbook_path.exists():
        logger.error(f"Company workbook not found: {workbook_path}")
        return 1

    # Validate company workbook
    errors = validate_company_workbook(str(workbook_path))
    if errors:
        logger.error("Company workbook validation failed:")
        for err in errors:
            logger.error(f"  - {err}")
        return 1

    # Determine checklist path
    if args.checklist:
        checklist_path = Path(args.checklist).resolve()
    else:
        # Default: look for netnet_diligence_checklist.xlsx in project root
        # Try parent directories up to 3 levels
        checklist_path = None
        search_dir = workbook_path.parent
        for _ in range(4):
            candidate = search_dir / 'netnet_diligence_checklist.xlsx'
            if candidate.exists():
                checklist_path = candidate
                break
            search_dir = search_dir.parent

        if checklist_path is None:
            logger.error("Could not find netnet_diligence_checklist.xlsx. Use --checklist to specify path.")
            return 1

    if not checklist_path.exists():
        logger.error(f"Checklist template not found: {checklist_path}")
        return 1

    # Determine output path
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        # Default: create in same folder as company workbook
        company_name = workbook_path.stem
        output_path = workbook_path.parent / f"{company_name}_netnet_diligence_checklist.xlsx"

    logger.info(f"Company workbook: {workbook_path}")
    logger.info(f"Checklist template: {checklist_path}")
    logger.info(f"Output: {output_path}")

    # Extract data from company workbook
    logger.info("Extracting data from company workbook...")
    reader = CompanyWorkbookReader(str(workbook_path), args.price)
    data = reader.extract_all()
    reader.close()

    # Report-only mode: just print and exit
    if args.report_only:
        print_report(data, args.market_cap, args.price, {}, str(output_path))
        return 0

    # Dry-run mode: show what would be written
    if args.dry_run:
        print("\n[DRY RUN] Would write the following:")
        print_report(data, args.market_cap, args.price,
                    {'simulated': 'dry-run'}, str(output_path))
        return 0

    # Copy template to output location
    logger.info(f"Copying template to: {output_path}")
    shutil.copy2(str(checklist_path), str(output_path))

    # Write data to checklist
    logger.info("Writing data to Stage1_Quantitative sheet...")
    writer = Stage1Writer(str(output_path))
    populated = writer.populate(data, args.market_cap)
    writer.save()
    writer.close()

    # Print report
    print_report(data, args.market_cap, args.price, populated, str(output_path))

    return 0


if __name__ == '__main__':
    exit(main())
