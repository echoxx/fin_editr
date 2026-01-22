#!/usr/bin/env python3
"""
Populate the Overview tab in net-net analysis workbooks.

Auto-fills fields that can be determined from:
- Ticker symbol (user input)
- Exchange code (maps to country and currency)
- Company info scraped from Investing.com
"""

from pathlib import Path
from typing import Optional
from dataclasses import dataclass

import openpyxl


# Exchange code to country/currency mapping
EXCHANGE_INFO = {
    # Japan
    "TYO": ("Japan", "JPY", "Yen"),
    "TSE": ("Japan", "JPY", "Yen"),
    "TOKYO": ("Japan", "JPY", "Yen"),

    # Hong Kong
    "HKG": ("Hong Kong", "HKD", "HK Dollar"),
    "HKEX": ("Hong Kong", "HKD", "HK Dollar"),

    # United States
    "NYSE": ("United States", "USD", "US Dollar"),
    "NASDAQ": ("United States", "USD", "US Dollar"),
    "AMEX": ("United States", "USD", "US Dollar"),
    "OTC": ("United States", "USD", "US Dollar"),

    # United Kingdom
    "LON": ("United Kingdom", "GBP", "Pounds"),
    "LSE": ("United Kingdom", "GBP", "Pounds"),

    # Europe
    "EPA": ("France", "EUR", "Euro"),
    "EURONEXT": ("France", "EUR", "Euro"),
    "ETR": ("Germany", "EUR", "Euro"),
    "FRA": ("Germany", "EUR", "Euro"),
    "AMS": ("Netherlands", "EUR", "Euro"),
    "BRU": ("Belgium", "EUR", "Euro"),
    "MIL": ("Italy", "EUR", "Euro"),
    "BME": ("Spain", "EUR", "Euro"),

    # Other Asia
    "KRX": ("South Korea", "KRW", "Won"),
    "KOSDAQ": ("South Korea", "KRW", "Won"),
    "SGX": ("Singapore", "SGD", "Singapore Dollar"),
    "TWSE": ("Taiwan", "TWD", "Taiwan Dollar"),
    "TPE": ("Taiwan", "TWD", "Taiwan Dollar"),
    "ASX": ("Australia", "AUD", "Australian Dollar"),

    # Canada
    "TSX": ("Canada", "CAD", "Canadian Dollar"),
    "CVE": ("Canada", "CAD", "Canadian Dollar"),
}


@dataclass
class OverviewFields:
    """Fields that can be populated in the Overview tab."""
    ticker: Optional[str] = None
    country: Optional[str] = None
    currency: Optional[str] = None
    price: Optional[float] = None
    exchange_rate: Optional[float] = None
    website: Optional[str] = None


def detect_country_from_exchange(exchange_code: str) -> tuple[Optional[str], Optional[str]]:
    """
    Determine country and currency from exchange code.

    Args:
        exchange_code: Exchange identifier (e.g., "TYO", "NYSE")

    Returns:
        Tuple of (country, currency_name) or (None, None) if unknown
    """
    exchange_upper = exchange_code.upper().strip()

    info = EXCHANGE_INFO.get(exchange_upper)
    if info:
        return info[0], info[2]  # country, currency name

    return None, None


def populate_overview(
    workbook_path: str | Path,
    ticker: str = None,
    exchange: str = None,
    country: str = None,
    currency: str = None,
    price: float = None,
    exchange_rate: float = None,
    website: str = None,
) -> dict:
    """
    Populate Overview tab with available data.

    Args:
        workbook_path: Path to the workbook
        ticker: Stock ticker symbol
        exchange: Exchange code (used to infer country/currency if not provided)
        country: Country name (optional, inferred from exchange if not provided)
        currency: Currency name (optional, inferred from exchange if not provided)
        price: Stock price (optional)
        exchange_rate: Exchange rate to USD (optional)
        website: Company website URL (optional)

    Returns:
        Dict of fields that were populated
    """
    workbook_path = Path(workbook_path)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    # Infer country/currency from exchange if not provided
    if exchange and (not country or not currency):
        inferred_country, inferred_currency = detect_country_from_exchange(exchange)
        if not country and inferred_country:
            country = inferred_country
        if not currency and inferred_currency:
            currency = inferred_currency

    # Load workbook
    wb = openpyxl.load_workbook(str(workbook_path))

    if "Overview" not in wb.sheetnames:
        raise ValueError(f"Workbook has no 'Overview' sheet: {workbook_path}")

    ws = wb["Overview"]

    # Track what we populate
    populated = {}

    # Cell mapping based on the workbook structure:
    # C3: Ticker
    # C4: Country
    # C5: Currency
    # C6: Price (input)
    # C7: Exchange Rate
    # C14: Website

    if ticker:
        ws["C3"] = ticker
        populated["ticker"] = ticker
        print(f"  Set Ticker: {ticker}")

    if country:
        ws["C4"] = country
        populated["country"] = country
        print(f"  Set Country: {country}")

    if currency:
        ws["C5"] = currency
        populated["currency"] = currency
        print(f"  Set Currency: {currency}")

    if price is not None:
        ws["C6"] = price
        populated["price"] = price
        print(f"  Set Price: {price}")

    if exchange_rate is not None:
        ws["C7"] = exchange_rate
        populated["exchange_rate"] = exchange_rate
        print(f"  Set Exchange Rate: {exchange_rate}")

    if website:
        ws["C14"] = website
        populated["website"] = website
        print(f"  Set Website: {website}")

    # Save workbook
    wb.save(str(workbook_path))
    wb.close()

    return populated


def clear_overview_manual_fields(workbook_path: str | Path) -> None:
    """
    Clear manual entry fields in Overview tab (for template preparation).

    Clears: Industry, Ticker, Country, Currency, Price, Exchange Rate,
            Market Cap, Date Founded, Date Listed, Website

    Preserves formula fields: Company, Shares Outstanding, NCAV/Share, P/NCAV

    Args:
        workbook_path: Path to the workbook
    """
    workbook_path = Path(workbook_path)

    wb = openpyxl.load_workbook(str(workbook_path))
    ws = wb["Overview"]

    # Clear manual fields (preserve formulas)
    manual_cells = ["C2", "C3", "C4", "C5", "C6", "C7", "C8", "C12", "C13", "C14"]

    for cell_ref in manual_cells:
        cell = ws[cell_ref]
        # Only clear if it's not a formula
        if cell.value and not str(cell.value).startswith("="):
            cell.value = None

    wb.save(str(workbook_path))
    wb.close()

    print(f"Cleared manual fields in Overview tab: {workbook_path}")


def read_overview_fields(workbook_path: str | Path) -> dict:
    """
    Read current values from Overview tab.

    Args:
        workbook_path: Path to the workbook

    Returns:
        Dict of field names to values
    """
    workbook_path = Path(workbook_path)

    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    ws = wb["Overview"]

    fields = {
        "company": ws["C1"].value,
        "industry": ws["C2"].value,
        "ticker": ws["C3"].value,
        "country": ws["C4"].value,
        "currency": ws["C5"].value,
        "price": ws["C6"].value,
        "exchange_rate": ws["C7"].value,
        "market_cap": ws["C8"].value,
        "shares_outstanding": ws["C9"].value,
        "ncav_per_share": ws["C10"].value,
        "p_ncav": ws["C11"].value,
        "date_founded": ws["C12"].value,
        "date_listed": ws["C13"].value,
        "website": ws["C14"].value,
    }

    wb.close()
    return fields


def main():
    """Test Overview populator functions."""
    import argparse

    parser = argparse.ArgumentParser(description="Populate Overview tab")
    parser.add_argument("workbook", help="Path to workbook")
    parser.add_argument("--ticker", "-t", help="Ticker symbol")
    parser.add_argument("--exchange", "-e", help="Exchange code (e.g., TYO)")
    parser.add_argument("--country", help="Country name")
    parser.add_argument("--currency", help="Currency name")
    parser.add_argument("--price", type=float, help="Stock price")
    parser.add_argument("--exchange-rate", type=float, help="Exchange rate to USD")
    parser.add_argument("--website", help="Company website")
    parser.add_argument("--read", action="store_true", help="Read current values")
    parser.add_argument("--clear", action="store_true", help="Clear manual fields")

    args = parser.parse_args()

    if args.read:
        fields = read_overview_fields(args.workbook)
        print("\nOverview tab fields:")
        for key, value in fields.items():
            print(f"  {key}: {value}")
        return

    if args.clear:
        clear_overview_manual_fields(args.workbook)
        return

    populated = populate_overview(
        args.workbook,
        ticker=args.ticker,
        exchange=args.exchange,
        country=args.country,
        currency=args.currency,
        price=args.price,
        exchange_rate=args.exchange_rate,
        website=args.website,
    )

    if populated:
        print(f"\nPopulated {len(populated)} field(s)")
    else:
        print("\nNo fields to populate")


if __name__ == "__main__":
    main()
